from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


FILE = "Клиент_Алексей_Иванов_ведение.xlsx"

COLORS = {
    "navy": "1F4E78",
    "dark_blue": "17365D",
    "blue": "D9EAF7",
    "light_blue": "EAF4FB",
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "purple": "EADCF8",
    "gray": "E7E6E6",
    "white": "FFFFFF",
    "black": "000000",
}

thin = Side(style="thin", color="D9D9D9")


def fill(color):
    return PatternFill("solid", fgColor=color)


def style(cell, bg=None, bold=False, color="000000", size=10, center=False):
    cell.font = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = fill(bg)
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )


def dropdown(ws, cell_range, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def rebuild_sheet(wb, name, index=None):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        del wb[name]
        return wb.create_sheet(name, idx)
    if index is None:
        return wb.create_sheet(name)
    return wb.create_sheet(name, index)


def build_table_sheet(ws, title, description, headers, widths, sample_rows=None):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title)
    style(ws.cell(1, 1), COLORS["navy"], True, COLORS["white"], 15, True)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1, description)
    style(ws.cell(2, 1), COLORS["gray"], False, COLORS["black"], 10, False)
    ws.row_dimensions[2].height = 34

    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
        style(ws.cell(4, col), COLORS["dark_blue"], True, COLORS["white"], 9, True)

    if sample_rows:
        for row_idx, row in enumerate(sample_rows, 5):
            for col, value in enumerate(row, 1):
                ws.cell(row_idx, col, value)
                bg = COLORS["light_blue"] if col <= len(headers) else COLORS["white"]
                style(ws.cell(row_idx, col), bg, False, COLORS["black"], 9, col in [2, 5, 7, 8])

    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}200"


def main():
    wb = load_workbook(FILE)

    insert_after = wb.sheetnames.index("Навигация тренировок") + 1 if "Навигация тренировок" in wb.sheetnames else len(wb.sheetnames)

    bot_clients = rebuild_sheet(wb, "Bot Clients", insert_after)
    build_table_sheet(
        bot_clients,
        "Bot Clients | Клиенты Telegram-бота",
        "Связывает клиента из таблицы с Telegram ID, временем уведомлений и активным тренировочным циклом.",
        [
            "client_id", "telegram_id", "client_name", "connect_code", "timezone", "morning_time",
            "measurement_day", "active_cycle", "active_week", "status", "coach_chat_id", "drive_folder_url", "notes",
        ],
        [18, 16, 22, 16, 18, 14, 18, 18, 12, 12, 16, 30, 30],
        [[
            "alexey_ivanov", "", "Алексей Иванов", "ALEXEY2026", "Europe/Kyiv", "07:00",
            "Понедельник", "Гипертрофия", 1, "active", "", "Вставить ссылку", "Тестовый клиент MVP",
        ]],
    )
    dropdown(bot_clients, "H5:H200", ["Гипертрофия", "Интервальная", "Выносливость"])
    dropdown(bot_clients, "J5:J200", ["active", "paused", "archived"])

    bot_state = rebuild_sheet(wb, "Bot State")
    build_table_sheet(
        bot_state,
        "Bot State | Текущее состояние диалога",
        "Нужно, чтобы бот понимал, какой следующий ответ ожидается от клиента.",
        [
            "client_id", "telegram_id", "action", "step", "cycle", "week", "day_title", "exercise_index",
            "exercise_name", "temp_sets", "temp_reps", "temp_weight", "temp_rpe", "temp_comment", "updated_at",
        ],
        [18, 16, 24, 24, 16, 10, 28, 14, 28, 12, 18, 14, 12, 30, 20],
    )

    bot_schedule = rebuild_sheet(wb, "Bot Schedule")
    build_table_sheet(
        bot_schedule,
        "Bot Schedule | Расписание сообщений",
        "Очередь напоминаний: тренировки, замеры, фото, чек-ины и повторные уведомления.",
        ["schedule_id", "client_id", "type", "date", "time", "message_key", "status", "sent_at", "error", "payload"],
        [20, 18, 16, 13, 10, 22, 12, 20, 30, 40],
        [["sch_001", "alexey_ivanov", "workout", date(2026, 6, 1), "07:00", "today_workout", "pending", "", "", "Гип W1"]],
    )
    dropdown(bot_schedule, "C5:C500", ["workout", "measurements", "photos", "checkin", "reminder"])
    dropdown(bot_schedule, "G5:G500", ["pending", "sent", "skipped", "failed"])

    raw_results = rebuild_sheet(wb, "Exercise Results Raw")
    build_table_sheet(
        raw_results,
        "Exercise Results Raw | Сырые результаты упражнений",
        "Главная база фактических тренировочных данных для аналитики и отчетов. Бот добавляет сюда каждое выполненное упражнение.",
        [
            "datetime", "client_id", "cycle", "week", "day_title", "exercise_order", "exercise_name",
            "planned_sets", "planned_reps", "planned_weight", "actual_sets", "actual_reps", "actual_weight",
            "actual_rpe", "comment", "source_sheet", "source_row",
        ],
        [20, 18, 16, 9, 28, 14, 28, 12, 14, 14, 12, 18, 14, 10, 34, 14, 10],
    )

    photo_uploads = rebuild_sheet(wb, "Photo Uploads")
    build_table_sheet(
        photo_uploads,
        "Photo Uploads | Загруженные фото",
        "Сюда бот пишет все фото и ссылки на Google Drive. Лист 'Фото и состав тела' остается красивым клиентским видом.",
        [
            "datetime", "client_id", "week", "photo_type", "telegram_file_id", "drive_file_url",
            "drive_folder_url", "target_sheet", "target_cell", "status", "notes",
        ],
        [20, 18, 9, 16, 34, 34, 34, 18, 12, 12, 30],
    )
    dropdown(photo_uploads, "D5:D500", ["front", "side", "back", "body_composition", "other"])
    dropdown(photo_uploads, "J5:J500", ["uploaded", "failed", "pending"])

    checkins = rebuild_sheet(wb, "Check-ins")
    build_table_sheet(
        checkins,
        "Check-ins | Еженедельные ответы клиента",
        "Короткий недельный отчет клиента: adherence, сложности, энергия, вопросы тренеру.",
        [
            "date", "client_id", "week", "adherence", "energy", "stress", "sleep_quality", "main_win",
            "main_problem", "questions", "coach_summary", "status",
        ],
        [13, 18, 9, 12, 10, 10, 14, 34, 34, 34, 34, 12],
    )
    dropdown(checkins, "L5:L500", ["new", "reviewed", "answered"])

    alerts = rebuild_sheet(wb, "Alerts")
    build_table_sheet(
        alerts,
        "Alerts | Флаги внимания",
        "Отдельный лист для рисков: высокая усталость, боль, пропуски, низкий сон, очень высокий RPE.",
        ["datetime", "client_id", "type", "severity", "message", "source", "status", "coach_note"],
        [20, 18, 18, 12, 44, 20, 12, 34],
    )
    dropdown(alerts, "D5:D500", ["low", "medium", "high"])
    dropdown(alerts, "G5:G500", ["new", "seen", "resolved"])

    bot_logs = rebuild_sheet(wb, "Bot Logs")
    build_table_sheet(
        bot_logs,
        "Bot Logs | Журнал действий бота",
        "Технический журнал для отладки: что отправлено, что записано, где была ошибка.",
        ["datetime", "client_id", "telegram_id", "action", "status", "details", "raw_payload"],
        [20, 18, 16, 24, 12, 44, 60],
    )
    dropdown(bot_logs, "E5:E1000", ["success", "failed", "skipped", "pending"])

    # Keep business sheets before training plans, then technical bot sheets, then plans.
    preferred = [
        "Инструкция", "Профиль клиента", "Прогресс тела", "Фото и состав тела", "Самочувствие",
        "Дашборд", "База упражнений", "Навигация тренировок", "Bot Clients", "Bot State",
        "Bot Schedule", "Exercise Results Raw", "Photo Uploads", "Check-ins", "Alerts", "Bot Logs",
    ]
    training = [f"Гип W{i}" for i in range(1, 9)] + [f"Инт W{i}" for i in range(1, 9)]
    order = [name for name in preferred + training if name in wb.sheetnames]
    others = [name for name in wb.sheetnames if name not in order]
    wb._sheets = [wb[name] for name in order + others]

    wb.active = wb.sheetnames.index("Bot Clients")
    wb.save(FILE)


if __name__ == "__main__":
    main()
