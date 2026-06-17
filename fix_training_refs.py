from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


FILE = "Клиент_Алексей_Иванов_ведение.xlsx"

thin = Side(style="thin", color="D9D9D9")


def fill(color):
    return PatternFill("solid", fgColor=color)


wb = load_workbook(FILE)

if "Инструкция" in wb.sheetnames:
    ws = wb["Инструкция"]
    replacements = {
        "1. Тренер заполняет страницу 'План недели'.": "1. Тренер заполняет недельные листы: 'Гип W1...W8' или 'Инт W1...W8'.",
        "2. Клиент после тренировки заполняет 'Журнал тренировок': факт веса, повторов, RPE и комментарий.": "2. Клиент после каждого упражнения заполняет факт прямо в этом же недельном листе: подходы, повторы, вес, RPE, боль и комментарий.",
    }
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in replacements:
                cell.value = replacements[cell.value]

if "Дашборд" in wb.sheetnames:
    ws = wb["Дашборд"]
    ws["B10"] = "Смотреть недельные листы"
    ws["B10"].fill = fill("E7E6E6")
    ws["B10"].border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws["B10"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["D4"] = "Тренировочный план и журнал объединены в недельные листы. Для работы открой 'Навигация тренировок' и перейди к нужной неделе."

wb.save(FILE)
