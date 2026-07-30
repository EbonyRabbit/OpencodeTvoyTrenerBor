import shutil
from datetime import datetime, timedelta
from copy import copy

import openpyxl

TEMPLATE = "Шаблон_hyrox_ведение.xlsx"
OUTPUT = "Сила_Новичка_12_недель.xlsx"

PROGRAM_START = datetime(2026, 6, 1)
CYCLE_NAME = "СИЛА НОВИЧКА"
RUSSIAN_DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

# Прогрессия: неделя -> (вес, rpe)
WEEKLY_PARAMS = {
    1:  ("60%",  "6-7"),   2:  ("60%",  "6-7"),
    3:  ("62.5%","7"),     4:  ("50%",  "6"),
    5:  ("65%",  "7-8"),   6:  ("65%",  "7-8"),
    7:  ("67.5%","8"),     8:  ("55%",  "6"),
    9:  ("70%",  "8"),     10: ("70%",  "8"),
    11: ("72.5%","8-9"),   12: ("72.5%","8-9"),
}

WEEKLY_FOCUS = {
    1:  "Адаптация: техника, объём",
    2:  "Адаптация: прогрессия",
    3:  "Рост нагрузки",
    4:  "⚠️ РАЗГРУЗКА: -30% объёма",
    5:  "Развитие: сила",
    6:  "Развитие: прогрессия",
    7:  "Рост интенсивности",
    8:  "⚠️ РАЗГРУЗКА: восстановление",
    9:  "Пик: максимальная нагрузка",
    10: "Пик: прогрессия",
    11: "Финальная неделя",
    12: "Тест: проверка результатов",
}

DELOAD_WEEKS = {4, 8}


def get_sets_rpe(week):
    weight, rpe = WEEKLY_PARAMS[week]
    is_deload = week in DELOAD_WEEKS
    if is_deload:
        return "2", "6"
    if week <= 2:
        return "3", rpe
    if week <= 4:
        return "3", rpe
    if week <= 8:
        return "4", rpe
    return "4", rpe


def day_a_exercises(week):
    """День 1: Ноги + Жим + Тяга"""
    sets, rpe = get_sets_rpe(week)
    weight, _ = WEEKLY_PARAMS[week]
    is_deload = week in DELOAD_WEEKS

    base_sets = "2" if is_deload else sets
    hyper_sets = "2" if is_deload else "3"

    return [
        ("Разминка", "Суставная разминка + активация", "1", "10 мин", "", "5-6", "", ""),
        ("База", "Жим ногами в тренажёре", base_sets, "12-15", weight, rpe, "2010", "90 с"),
        ("База", "Жим штанги лёжа", base_sets, "10-12", weight, rpe, "2010", "90 с"),
        ("База", "Тяга горизонтальная в тренажёре", base_sets, "10-12", weight, rpe, "2010", "90 с"),
        ("Гипертрофия", "Жим вверх гантели сидя", hyper_sets, "10-12", "рабочий", "7", "2011", "60 с"),
        ("Гипертрофия", "Тяга верхнего блока шире плеч хватом", hyper_sets, "10-12", "рабочий", "7", "2010", "60 с"),
        ("Гипертрофия", "Сгибание голени лёжа/сидя", hyper_sets, "12-15", "рабочий", "7", "2011", "60 с"),
        ("Гипертрофия", "Планка прямая и боковая", hyper_sets, "30-45 с", "", "7", "", "60 с"),
        ("Заминка", "Растяжка всего тела", "1", "10 мин", "", "", "", ""),
    ]


def day_b_exercises(week):
    """День 2: Выпады + Жим + Тяга"""
    sets, rpe = get_sets_rpe(week)
    weight, _ = WEEKLY_PARAMS[week]
    is_deload = week in DELOAD_WEEKS

    base_sets = "2" if is_deload else sets
    hyper_sets = "2" if is_deload else "3"

    return [
        ("Разминка", "Суставная разминка + активация", "1", "10 мин", "", "5-6", "", ""),
        ("База", "Выпады шагающие с гантелями", base_sets, "10-12/н", weight, rpe, "2011", "90 с"),
        ("База", "Жим гантелей на наклонной", base_sets, "10-12", weight, rpe, "2010", "90 с"),
        ("База", "Тяга 1 гантели в упоре на скамью", base_sets, "10-12/р", weight, rpe, "2010", "90 с"),
        ("Гипертрофия", "Жим штанги стоя", hyper_sets, "10-12", "рабочий", "7", "2010", "60 с"),
        ("Гипертрофия", "Тяга верхнего блока обратным хватом", hyper_sets, "10-12", "рабочий", "7", "2010", "60 с"),
        ("Гипертрофия", "Экстензия", hyper_sets, "12-15", "рабочий", "7", "2011", "60 с"),
        ("Гипертрофия", "Пресс кранчи", hyper_sets, "15-20", "свой вес", "7", "", "45 с"),
        ("Заминка", "Растяжка всего тела", "1", "10 мин", "", "", "", ""),
    ]


def day_c_exercises(week):
    """День 3: Гоблет + Отжимания + Тяга к лицу"""
    sets, rpe = get_sets_rpe(week)
    weight, _ = WEEKLY_PARAMS[week]
    is_deload = week in DELOAD_WEEKS

    base_sets = "2" if is_deload else sets
    hyper_sets = "2" if is_deload else "3"

    return [
        ("Разминка", "Суставная разминка + активация", "1", "10 мин", "", "5-6", "", ""),
        ("База", "Гоблет-приседания", base_sets, "10-12", weight, rpe, "2010", "90 с"),
        ("База", "Отжимания", base_sets, "10-15", "свой вес", rpe, "2010", "90 с"),
        ("База", "Тяга широким хватом к лицу стоя", base_sets, "10-12", weight, rpe, "2010", "90 с"),
        ("Гипертрофия", "Подтягивания в гравитроне", hyper_sets, "8-12", "обратный вес", "7", "2010", "60 с"),
        ("Гипертрофия", "Отведение плеч в сторону", hyper_sets, "12-15", "лёгкий", "7", "2011", "60 с"),
        ("Гипертрофия", "Боковые выпады", hyper_sets, "10-12/н", "рабочий", "7", "2011", "60 с"),
        ("Гипертрофия", "Обратные скручивания", hyper_sets, "15-20", "свой вес", "7", "", "45 с"),
        ("Заминка", "Растяжка всего тела", "1", "10 мин", "", "", "", ""),
    ]


DAY_NAMES = ["День 1", "День 2", "День 3"]
DAY_GENERATORS = [day_a_exercises, day_b_exercises, day_c_exercises]


def is_merged_cell(ws, row, col):
    """Проверяет, является ли ячейка частью объединённого диапазона (не первой в диапазоне)."""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            # Это объединённая ячейка
            if row == merged_range.min_row and col == merged_range.min_col:
                return False  # Это первая ячейка — можно писать
            return True  # Это не первая ячейка — нельзя писать
    return False


def clear_rows(ws, start_row, count):
    """Очищает значения во всех ячейках строки (кроме объединённых не-первых)."""
    for i in range(count):
        row = start_row + i
        for col in range(1, 14):
            if not is_merged_cell(ws, row, col):
                ws.cell(row, col).value = None


def write_exercises(ws, start_row, exercises):
    """Записывает упражнения начиная с start_row."""
    for i, ex in enumerate(exercises):
        row = start_row + i
        for col in range(1, 9):
            if not is_merged_cell(ws, row, col):
                ws.cell(row, col).value = ex[col - 1] if col - 1 < len(ex) else None


def process_week(ws, week_num):
    """Обрабатывает один недельный лист, заменяя данные упражнений."""
    focus = WEEKLY_FOCUS[week_num]
    is_deload = week_num in DELOAD_WEEKS
    weight, rpe = WEEKLY_PARAMS[week_num]

    # Обновить заголовок (строка 1)
    ws.cell(1, 1).value = f"{CYCLE_NAME} | Неделя {week_num}"

    # Обновить подзаголовок (строка 2)
    ws.cell(2, 1).value = f"Период: Неделя {week_num} | {focus} | Базовый вес: {weight}"

    # Для W1-W8: 4 тренировочных дня
    # День 1: R8-R16 (9 строк данных), R5 заголовок, R6 цель, R7 колонки
    # День 2: R21-R29, R18 заголовок, R19 цель, R20 колонки
    # День 3: R34-R42, R31 заголовок, R32 цель, R33 колонки
    # День 4: R46-R48, R43 заголовок, R44 цель, R45 колонки

    # Определяем структуру по неделям
    if week_num <= 8:
        # W1-W8: 48 строк
        days = [
            (5, 8, 9, "Понедельник", DAY_NAMES[0]),
            (18, 21, 9, "Среда", DAY_NAMES[1]),
            (31, 34, 9, "Пятница", DAY_NAMES[2]),
            (43, 46, 3, "Суббота", "Восстановление"),
        ]
    elif week_num <= 11:
        # W9-W11: 45 строк
        days = [
            (5, 8, 9, "Понедельник", DAY_NAMES[0]),
            (17, 20, 9, "Среда", DAY_NAMES[1]),
            (29, 32, 7, "Пятница", DAY_NAMES[2]),
            (40, 43, 3, "Суббота", "Восстановление"),
        ]
    else:
        # W12: особая структура — тестовая неделя
        process_week12(ws, week_num)
        return

    for day_idx, (header_row, data_start, max_rows, day_name, day_type) in enumerate(days):
        # Заголовок дня
        ws.cell(header_row, 1).value = f"{day_name} | {day_type}"

        # Цель
        if day_idx < 3:
            ws.cell(header_row + 1, 1).value = f"🎯 Неделя {week_num}: {focus.replace('⚠️ ', '')}"
        else:
            ws.cell(header_row + 1, 1).value = "🎯 Восстановление"

        # Упражнения
        if day_idx < 3:
            exercises = DAY_GENERATORS[day_idx](week_num)
            clear_rows(ws, data_start, max_rows)
            write_exercises(ws, data_start, exercises)
        else:
            # День 4 — восстановление / лёгкий бег
            recovery = [
                ("Разминка", "Ходьба + динамическая растяжка + ускорения", "1", "10 мин", "", "5-6", "", ""),
                ("Бег", "Лёгкий бег", "1", "3-4 км", "ЧСС 120-135", "Носовое дыхание", "", ""),
                ("Заминка", "Ходьба + растяжка", "1", "10 мин", "", "", "", ""),
            ]
            clear_rows(ws, data_start, max_rows)
            write_exercises(ws, data_start, recovery)


def process_week12(ws, week_num):
    """Тестовая неделя — особая структура."""
    ws.cell(2, 1).value = "Период: Неделя 12 | Тест: проверка результатов"

    # День 1: разминка + разогрев
    ws.cell(5, 1).value = "Понедельник | Разминка"
    ws.cell(6, 1).value = "🎯 Неделя 12: Тест — подготовка"

    warmup = [
        ("Разминка", "Лёгкая мобилизация + растяжка", "1", "15 мин", "", "", "", ""),
        ("Активация", "Жим ногами (лёгкий вес)", "2", "12", "60%", "6", "", "—"),
        ("Активация", "Тяга горизонтальная (лёгкий)", "2", "12", "60%", "6", "", "—"),
        ("Активация", "Жим штанги лёжа (лёгкий)", "2", "12", "60%", "6", "", "—"),
        ("Активация", "Лёгкий бег 1 км", "1", "1 км", "", "6", "", "—"),
        ("Заминка", "Растяжка + фоам-ролл", "1", "15 мин", "", "", "", ""),
    ]
    clear_rows(ws, 8, 6)
    write_exercises(ws, 8, warmup)

    # День 2: тест всех упражнений
    ws.cell(15, 1).value = "Среда | Тест"
    ws.cell(16, 1).value = "🎯 Максимальные повторы на весе 70%"

    test_day = [
        ("Разминка", "Суставная разминка + активация", "1", "15 мин", "", "5-6", "", ""),
        ("Тест", "Жим штанги лёжа — тест на макс. повторы", "1", "AMRAP", "70%", "max", "", "3 мин"),
        ("Тест", "Тяга горизонтальная — тест на макс. повторы", "1", "AMRAP", "70%", "max", "", "3 мин"),
        ("Тест", "Жим вверх гантели сидя — тест на макс. повторы", "1", "AMRAP", "70%", "max", "", "3 мин"),
        ("Тест", "Жим ногами — тест на макс. повторы", "1", "AMRAP", "70%", "max", "", "3 мин"),
        ("Заминка", "Растяжка всего тела", "1", "15 мин", "", "", "", ""),
    ]
    clear_rows(ws, 17, 19)
    write_exercises(ws, 17, test_day)

    # День 3: восстановление
    ws.cell(29, 1).value = "Пятница | Восстановление"
    ws.cell(30, 1).value = "🎯 Полное восстановление после тестов"

    recovery = [
        ("Восстановление", "Полный отдых / активная ходьба", "", "", "", "", "", ""),
        ("Восстановление", "Лёгкая растяжка + фоам-ролл", "1", "20 мин", "", "", "", ""),
        ("Восстановление", "Лёгкий бег 2-3 км", "1", "2-3 км", "ЧСС 100-120", "", "", "без боли"),
    ]
    clear_rows(ws, 31, 7)
    write_exercises(ws, 31, recovery)

    # День 4: ходьба
    ws.cell(40, 1).value = "Суббота | Восстановление"
    ws.cell(41, 1).value = "🎯 Активное восстановление"

    walk = [
        ("Восстановление", "Ходьба 3-5 км в лёгком темпе", "1", "3-5 км", "", "", "", ""),
        ("Восстановление", "Лёгкая растяжка", "1", "15 мин", "", "", "", ""),
    ]
    clear_rows(ws, 42, 3)
    write_exercises(ws, 42, walk)


def main():
    # Копируем шаблон
    shutil.copy2(TEMPLATE, OUTPUT)

    # Открываем копию
    wb = openpyxl.load_workbook(OUTPUT)

    # Обрабатываем только недельные листы
    for week_num in range(1, 13):
        sheet_name = f"W{week_num}"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            process_week(ws, week_num)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
