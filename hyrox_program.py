from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "HYROX_12_week_program.xlsx"
PROGRAM_START = datetime(2026, 6, 1)
CYCLE_NAME = "HYROX"
RUSSIAN_DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

NAVY = "1F4E78"
NAVY_DARK = "17365D"
WHITE = "FFFFFF"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF4FB"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
ORANGE = "FCE5CD"
LIGHT_GRAY = "E7E6E6"
RED_LIGHT = "F4CCCC"
PURPLE_LIGHT = "EADCF8"

thin = Side(style="thin", color="D9D9D9")

COL_WIDTHS = [14, 34, 11, 13, 18, 12, 14, 12, 11, 11, 11, 11, 26]

PLAN_HEADERS = ["Раздел", "Упражнение", "План подходы", "План повторы", "План вес", "RPE/RIR", "Темп", "Отдых"]
FACT_HEADERS = ["Факт подходы", "Факт повторы", "Факт вес", "Факт RPE", "Комментарий клиента"]
ALL_HEADERS = PLAN_HEADERS + FACT_HEADERS  # 13 columns


def fill(color):
    return PatternFill("solid", fgColor=color)


def font(bold=False, size=9, color=WHITE):
    return Font(bold=bold, size=size, color=color)


def style_title(ws, cell, text):
    c = ws[cell]
    c.value = text
    c.font = Font(bold=True, size=15, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row):
    for c in range(1, 14):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, size=9, color=WHITE)
        if c <= 8:
            cell.fill = fill(NAVY)
        else:
            cell.fill = fill(NAVY_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def write_data_row(ws, row, values, color=LIGHT_BLUE):
    for c in range(1, 14):
        cell = ws.cell(row, c)
        if c <= len(values) and values[c - 1] is not None and values[c - 1] != "":
            cell.value = values[c - 1]
        cell.font = Font(size=9)
        cell.fill = fill(color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def merge_and_fill(ws, start_row, end_row, col, text, fill_color, font_obj):
    if start_row == end_row:
        cell = ws.cell(start_row, col)
        cell.value = text
        cell.font = font_obj
        cell.fill = fill(fill_color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    else:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        cell = ws.cell(start_row, col)
        cell.value = text
        cell.font = font_obj
        cell.fill = fill(fill_color)
        cell.alignment = Alignment(vertical="center", wrap_text=True, text_rotation=90)
        for r in range(start_row, end_row + 1):
            for c in range(1, 14):
                ws.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)


def day_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = Font(bold=True, size=12, color=WHITE)
    cell.fill = fill(NAVY_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, 14):
        ws.cell(row, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
    ws.row_dimensions[row].height = 22


def goal_row(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    cell = ws.cell(row, 1)
    cell.value = "🎯 " + text
    cell.font = Font(bold=True, size=10)
    cell.fill = fill(YELLOW)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    for c in range(1, 14):
        ws.cell(row, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)


# ======================
# EXERCISE DATA
# ======================

S = {}  # (section, exercise, plan_sets, plan_reps, plan_weight, rpe, tempo, rest)


def combo(warmup, strength, hypertrophy, endurance, cooldown):
    return [
        ("Разминка", warmup[0], warmup[1], "", "", "", "", ""),
        *[("База", e[0], e[1], e[2], e[3], e[4], e[5]) for e in strength],
        *[("Гипертрофия", e[0], e[1], e[2], e[3], e[4], e[5]) for e in hypertrophy],
        *[("Выносливость", e[0], e[1], e[2], e[3], e[4], e[5]) for e in endurance],
        ("Заминка", cooldown[0], cooldown[1], "", "", "", "", ""),
    ]


def run_ex(km, pace, notes):
    return [
        ("Разминка", "Ходьба + динамическая растяжка + ускорения", "1", "10 мин", "", "5-6", "", ""),
        ("Бег", "Бег", "1", km, pace, notes, "", ""),
        ("Заминка", "Ходьба + растяжка", "1", "10 мин", "", "", "", ""),
    ]


p1 = {"main_sets": "4", "main_reps": "8-10", "main_rpe": "7", "main_tempo": "2010"}
p2 = {"main_sets": "4", "main_reps": "6-8", "main_rpe": "8", "main_tempo": "2010"}
p3 = {"main_sets": "5", "main_reps": "5-6", "main_rpe": "9", "main_tempo": "2010"}

# Week 1-4: Foundation
S["1_1"] = combo(
    ("Мобилизация + активация кора", "10 мин", "", "", "", "", ""),
    [
        ("Приседания со штангой", p1["main_sets"], p1["main_reps"], "70%", p1["main_rpe"], p1["main_tempo"], "90 с"),
        ("Жим штанги лёжа", p1["main_sets"], p1["main_reps"], "70%", p1["main_rpe"], p1["main_tempo"], "90 с"),
        ("Тяга штанги в наклоне", p1["main_sets"], p1["main_reps"], "70%", p1["main_rpe"], p1["main_tempo"], "90 с"),
    ],
    [
        ("Жим гантелей сидя", "3", "12-15", "рабочий", "7-8", "2011", "60 с"),
        ("Подтягивания", "3", "10-12", "свой вес", "7-8", "2010", "60 с"),
        ("Румынская тяга", "3", "12-15", "70%", "7", "2011", "60 с"),
    ],
    [
        ("AMRAP 10 мин: 500м гребля + 10 wall ball (6 кг)", "AMRAP", "10 мин", "ЧСС 145-160", "7-8", "", "—"),
    ],
    ("Растяжка грудного отдела + бёдер", "10 мин", "", "", "", "", ""),
)

S["1_2"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Фронтальные приседания", p1["main_sets"], p1["main_reps"], "70%", p1["main_rpe"], p1["main_tempo"], "90 с"),
        ("Тяга верхнего блока", p1["main_sets"], p1["main_reps"], "рабочий", p1["main_rpe"], p1["main_tempo"], "90 с"),
        ("Жим гантелей на наклонной", p1["main_sets"], p1["main_reps"], "рабочий", p1["main_rpe"], p1["main_tempo"], "90 с"),
    ],
    [
        ("Болгарские выпады", "3", "10-12/н", "рабочий", "7-8", "2011", "60 с"),
        ("Тяга гантели в наклоне", "3", "10-12/р", "рабочий", "7-8", "2010", "60 с"),
        ("Планка с отягощением", "3", "45-60 с", "+10 кг", "7", "", "60 с"),
    ],
    [
        ("4 раунда: 300м гребля + 10 бёрпи + 100м фермерская (24 кг)", "4", "раунда", "ЧСС 150-165", "7-8", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "10 мин", "", "", "", "", ""),
)

S["1_3"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Становая тяга", "4", "8-10", "70%", "7", "2010", "90 с"),
        ("Жим штанги стоя", "4", "8-10", "70%", "7", "2010", "90 с"),
    ],
    [
        ("Выпады назад с гантелями", "3", "10-12/н", "рабочий", "7-8", "2011", "60 с"),
        ("Тяга лица (face pull)", "3", "15-20", "рабочий", "7", "", "45 с"),
        ("Подъём ног в висе", "3", "12-15", "свой вес", "7", "", "60 с"),
    ],
    [
        ("5 раундов: 200м бег + 10 wall ball (6 кг) + 5 бёрпи", "5", "раундов", "ЧСС 155-170", "7-8", "", "—"),
    ],
    ("Растяжка бёдер + грудного отдела", "10 мин", "", "", "", "", ""),
)

S["1_run"] = run_ex("4-5 км", "ЧСС 120-135", "Носовое дыхание")
S["1_run_w2"] = run_ex("5 км", "ЧСС 120-135", "")
S["1_run_w3"] = run_ex("6 км", "ЧСС 120-135", "Прогрессия +1 км")
S["1_run_w4"] = run_ex("4 км", "ЧСС 110-125", "Разгрузка")

# Week 5-8: Development
S["2_1"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Приседания со штангой", "4", "6-8", "80%", "8", "2010", "120 с"),
        ("Жим штанги лёжа", "4", "6-8", "80%", "8", "2010", "120 с"),
        ("Тяга штанги в наклоне", "4", "6-8", "80%", "8", "2010", "120 с"),
    ],
    [
        ("Жим гантелей на наклонной", "3", "10-12", "рабочий", "8", "2011", "60 с"),
        ("Подтягивания с весом", "3", "8-10", "+10 кг", "8", "2010", "60 с"),
        ("Румынская тяга", "3", "10-12", "75%", "8", "2011", "60 с"),
    ],
    [
        ("500м гребля + 15 wall ball (9 кг) + 200м бег", "4", "раунда", "ЧСС 160-175", "8", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "10 мин", "", "", "", "", ""),
)

S["2_2"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Фронтальные приседания", "4", "6-8", "80%", "8", "2010", "120 с"),
        ("Жим гантелей стоя", "4", "8-10", "рабочий", "8", "2010", "90 с"),
        ("Тяга гантели в наклоне", "4", "8-10/р", "рабочий", "8", "2010", "90 с"),
    ],
    [
        ("Болгарские выпады", "3", "10-12/н", "рабочий", "8", "2011", "60 с"),
        ("Тяга верхнего блока широким хватом", "3", "12-15", "рабочий", "7-8", "2010", "60 с"),
        ("Планка + подъём ног", "3", "60 с + 15", "", "7", "", "45 с"),
    ],
    [
        ("400м бег + 20 wall ball (9 кг) + 50м санки толкать", "4", "раунда", "ЧСС 165-180", "8", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "10 мин", "", "", "", "", ""),
)

S["2_3"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Становая тяга", "4", "6-8", "80%", "8", "2010", "120 с"),
        ("Жим штанги лёжа", "4", "6-8", "80%", "8", "2010", "120 с"),
    ],
    [
        ("Выпады назад с гантелями", "3", "10-12/н", "рабочий", "8", "2011", "60 с"),
        ("Жим гантелей сидя", "3", "12-15", "рабочий", "8", "2011", "60 с"),
        ("Тяга лица (face pull)", "3", "15-20", "рабочий", "7", "", "45 с"),
    ],
    [
        ("AMRAP 15: 300м гребля + 12 wall ball (9 кг) + 8 бёрпи + 100м фермерская (28 кг)", "AMRAP 15", "мин", "ЧСС 165-180", "8", "", "—"),
    ],
    ("Растяжка всего тела", "10 мин", "", "", "", "", ""),
)

S["2_run_w5"] = run_ex("7 км", "ЧСС 120-135 + финиш 500м", "4×100м ускорения после")
S["2_run_w6"] = run_ex("8 км", "ЧСС 120-135, последние 2 км быстрее", "Прогрессивный бег")
S["2_run_w7"] = run_ex("8 км", "ЧСС 120-135", "")
S["2_run_w8"] = run_ex("5 км", "ЧСС 110-125", "Разгрузка")

# Week 9-11: Peak
S["3_1"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Приседания со штангой", "5", "5-6", "85%", "9", "2010", "120 с"),
        ("Жим штанги лёжа", "5", "5-6", "85%", "9", "2010", "120 с"),
        ("Тяга штанги в наклоне", "4", "6-8", "80%", "8", "2010", "120 с"),
    ],
    [
        ("Жим гантелей на наклонной", "3", "8-10", "рабочий", "8", "2011", "60 с"),
        ("Подтягивания с весом", "3", "6-8", "+15 кг", "8-9", "2010", "60 с"),
    ],
    [
        ("Полный Hyrox: 500м бег + 30 wall ball (9 кг) + 500м гребля + 20 бёрпи-прыжки", "AMRAP 15", "мин", "ЧСС 175-190", "9", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "15 мин", "", "", "", "", ""),
)

S["3_2"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Фронтальные приседания", "5", "5-6", "85%", "9", "2010", "120 с"),
        ("Жим гантелей стоя", "4", "6-8", "рабочий", "8-9", "2010", "90 с"),
    ],
    [
        ("Тяга гантели в наклоне", "4", "8-10/р", "рабочий", "8", "2010", "60 с"),
        ("Болгарские выпады", "3", "8-10/н", "рабочий", "8", "2011", "60 с"),
        ("Подъём ног в висе", "3", "15-20", "", "8", "", "45 с"),
    ],
    [
        ("40м санки толкать (тяж) + 40м тянуть (тяж) + 40м фермерская (32 кг)", "4", "раунда", "ЧСС 175-190", "9", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "15 мин", "", "", "", "", ""),
)

S["3_3"] = combo(
    ("Мобилизация + активация", "10 мин", "", "", "", "", ""),
    [
        ("Становая тяга", "4", "4-6", "85%", "9", "2010", "120 с"),
        ("Жим штанги лёжа", "4", "5-6", "85%", "9", "2010", "120 с"),
    ],
    [
        ("Выпады с sandbag (20 кг)", "3", "20 шагов", "20 кг", "8-9", "", "60 с"),
        ("Жим гантелей сидя", "3", "10-12", "рабочий", "8", "2011", "60 с"),
    ],
    [
        ("EMOM 10: 15 wall ball (9 кг) + 100м бег + 5 бёрпи", "EMOM 10", "мин", "ЧСС 175-190", "9", "", "—"),
    ],
    ("Растяжка + фоам-ролл", "10 мин", "", "", "", "", ""),
)

S["3_run_w9"] = run_ex("8 км", "10′ Zone2 + 15′ Zone3 + 10′ Zone2", "Прогрессивный бег")
S["3_run_w10"] = run_ex("9 км", "5 км Zone2 + 4 км Zone3", "Прогрессия")
S["3_run_w11"] = run_ex("5 км", "ЧСС 120-130, лёгкий", "5×100м ускорения после")

# Week 12: Race
S["12_1"] = [
    ("Разминка", "Лёгкая мобилизация + растяжка", "1", "15 мин", "", "", "", ""),
    ("Активация", "Wall ball (6 кг)", "3", "10", "6 кг", "6", "", "—"),
    ("Активация", "Гоблет-присед (16 кг)", "3", "8", "16 кг", "6", "", "—"),
    ("Активация", "Бёрпи", "3", "5", "", "6", "", "—"),
    ("Активация", "Лёгкий бег 1 км", "1", "1 км", "", "6", "", "—"),
    ("Заминка", "Растяжка + фоам-ролл", "1", "15 мин", "", "", "", ""),
]

S["12_2"] = [
    ("Разминка", "Лёгкий бег 1 км + мобилизация", "1", "15 мин", "", "", "", ""),
    ("Гонка", "SkiErg 1000 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Санки толкать 50 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Санки тянуть 50 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Burpee broad jumps 80 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Гребля 1000 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Фермерская прогулка 200 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Выпады с sandbag 100 м", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Гонка", "Wall ball 100", "", "", "RACE", "", "", ""),
    ("Бег", "Бег 1 км", "", "", "RACE", "", "", ""),
    ("Восстановление", "Ходьба + растяжка", "", "", "", "", "", "не ограничено"),
]

S["12_3"] = [
    ("Восстановление", "Полный отдых / активная ходьба", "", "", "", "", "", ""),
    ("Восстановление", "Лёгкая растяжка + фоам-ролл", "1", "20 мин", "", "", "", ""),
    ("Восстановление", "Лёгкий бег 2-3 км", "1", "2-3 км", "ЧСС 100-120", "", "", "без боли"),
]

S["12_run"] = [
    ("Восстановление", "Ходьба 3-5 км в лёгком темпе", "1", "3-5 км", "", "", "", ""),
]

WEEKLY = {
    1: ("Фундамент: техника, аэробная база", ["1_1", "1_2", "1_3"], "1_run"),
    2: ("Фундамент: увеличение объёма", ["1_1", "1_2", "1_3"], "1_run_w2"),
    3: ("Фундамент: рост интенсивности", ["1_1", "1_2", "1_3"], "1_run_w3"),
    4: ("⚠️ РАЗГРУЗКА: -30% объёма", ["1_1", "1_2", "1_3"], "1_run_w4"),
    5: ("Развитие: Hyrox-специфика", ["2_1", "2_2", "2_3"], "2_run_w5"),
    6: ("Развитие: рост интенсивности", ["2_1", "2_2", "2_3"], "2_run_w6"),
    7: ("Развитие: пиковая нагрузка", ["2_1", "2_2", "2_3"], "2_run_w7"),
    8: ("⚠️ РАЗГРУЗКА: восстановление", ["2_1", "2_2", "2_3"], "2_run_w8"),
    9: ("Пик: специфическая подготовка", ["3_1", "3_2", "3_3"], "3_run_w9"),
    10: ("Пик: макс. специфическая нагрузка", ["3_1", "3_2", "3_3"], "3_run_w10"),
    11: ("Предгоночная разгрузка", ["3_1", "3_2", "3_3"], "3_run_w11"),
    12: ("ГОНОЧНАЯ НЕДЕЛЯ: HYROX", ["12_1", "12_2", "12_3"], "12_run"),
}

DELOAD_WEEKS = {4, 8}
DAY_TYPES = ["Комбинированная 1", "Комбинированная 2", "Комбинированная 3", "Беговая работа"]
MERGE_COLORS = {
    "Разминка": LIGHT_BLUE, "Заминка": LIGHT_BLUE,
    "База": LIGHT_BLUE, "Гипертрофия": LIGHT_BLUE,
    "Выносливость": LIGHT_BLUE, "Бег": LIGHT_BLUE,
    "Гонка": LIGHT_BLUE, "Восстановление": LIGHT_BLUE,
    "Активация": LIGHT_BLUE, "Кор": LIGHT_BLUE,
}


def apply_deload(exs):
    out = []
    for ex in exs:
        s = list(ex)
        if s[0] not in ("Разминка", "Заминка", "Восстановление"):
            if s[2] and s[2] not in ("", "—") and s[2].isdigit():
                s[2] = str(max(1, int(s[2]) - 1))
            if s[5] and s[5] not in ("", "—"):
                s[5] = "6"
        out.append(s)
    return out


def write_day_block(ws, start_row, day_name, goal, exercises, deload=False):
    if deload:
        exercises = apply_deload(exercises)

    day_header(ws, start_row, day_name)
    r0 = start_row + 1

    if goal:
        goal_row(ws, r0, goal)
        r0 += 1

    for c, h in enumerate(ALL_HEADERS, 1):
        ws.cell(r0, c, h)
    style_header_row(ws, r0)
    r0 += 1

    for ex in exercises:
        row_vals = list(ex[:8]) + [""] * 5
        color = MERGE_COLORS.get(str(ex[0]), LIGHT_BLUE)
        write_data_row(ws, r0, row_vals, color)
        r0 += 1

    return r0 + 1


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    for week_num in range(1, 13):
        focus, sids, run_id = WEEKLY[week_num]
        is_deload = week_num in DELOAD_WEEKS
        ws = wb.create_sheet(f"W{week_num}")

        week_start = PROGRAM_START + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)
        period = f"{week_start.strftime('%d.%m.%Y')} - {week_end.strftime('%d.%m.%Y')}"

        ws.merge_cells("A1:M1")
        style_title(ws, "A1", f"{CYCLE_NAME} | Неделя {week_num}")

        ws.merge_cells("A2:M2")
        ws["A2"] = f"Период: {period} | {focus}"
        ws["A2"].font = Font(size=10)
        ws["A2"].fill = fill(LIGHT_GRAY)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 14):
            ws.cell(2, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)

        ws.merge_cells("A3:M3")
        ws["A3"] = "Заполнять после каждого упражнения: факт подходов, повторов, веса, RPE и короткий комментарий."
        ws["A3"].font = Font(size=10)
        ws["A3"].fill = fill(YELLOW)
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 14):
            ws.cell(3, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)

        prev_row = 5
        goals = [
            "Неделя " + str(week_num) + ": " + focus.replace("⚠️ ", ""),
            "Фокус: техника, контроль пульса, ровный темп",
            "Фокус: техника, контроль пульса, ровный темп",
            "Восстановительный бег, контроль пульса",
        ]

        for day_idx in range(4):
            d = week_start + timedelta(days=day_idx)
            date_str = d.strftime("%d.%m.%Y")
            day_name = RUSSIAN_DAYS[d.weekday()]
            is_run = day_idx == 3

            if is_run:
                sid = run_id
                title = f"{day_name} {date_str} | {DAY_TYPES[3]}"
            else:
                sid = sids[day_idx]
                title = f"{day_name} {date_str} | {DAY_TYPES[day_idx]}"

            ex_data = S[sid]
            prev_row = write_day_block(ws, prev_row, title, goals[day_idx], ex_data, is_deload)

        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    create_workbook()
