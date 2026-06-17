from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "Клиент_Алексей_Иванов_ведение.xlsx"
START_DATE = date(2026, 6, 1)


COLORS = {
    "navy": "1F4E78",
    "blue": "D9EAF7",
    "light_blue": "EAF4FB",
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "red": "F4CCCC",
    "purple": "EADCF8",
    "gray": "E7E6E6",
    "dark_gray": "666666",
    "white": "FFFFFF",
    "black": "000000",
}


thin = Side(style="thin", color="D9D9D9")
medium = Side(style="medium", color="9E9E9E")


def fill(color):
    return PatternFill("solid", fgColor=color)


def style_title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=16, color=COLORS["white"])
    ws[cell].fill = fill(COLORS["navy"])
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def style_section(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=12, color=COLORS["white"])
    ws[cell].fill = fill(COLORS["navy"])
    ws[cell].alignment = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, color=COLORS["white"])
        c.fill = fill(COLORS["navy"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def set_common(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def add_table(ws, name, ref):
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def add_dropdown(ws, cell_range, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def setup_instruction(wb):
    ws = wb.create_sheet("Инструкция")
    ws.merge_cells("A1:H1")
    style_title(ws, "A1", "Система ведения клиента: Алексей Иванов")
    ws.row_dimensions[1].height = 28
    rows = [
        ["Цвет", "Кто заполняет", "Что означает"],
        ["Голубой", "Клиент / тренер", "Основные поля ввода: фактический вес, повторы, самочувствие."],
        ["Фиолетовый", "Тренер", "Комментарии, решения, корректировки плана."],
        ["Серый", "Автоматически", "Формулы, расчеты, итоговые показатели. Не редактировать."],
        ["Желтый", "Внимание", "Показатель требует контроля: высокий RPE, стресс, недосып."],
        ["Красный", "Риск", "Боль, сильная усталость, низкое восстановление."],
    ]
    for r, row in enumerate(rows, 3):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_header_row(ws, 3, 1, 3)
    color_map = {
        4: COLORS["blue"],
        5: COLORS["purple"],
        6: COLORS["gray"],
        7: COLORS["yellow"],
        8: COLORS["red"],
    }
    for row, color in color_map.items():
        ws.cell(row, 1).fill = fill(color)

    ws["A11"] = "Как использовать файл"
    ws["A11"].font = Font(bold=True, size=13)
    usage = [
        "1. Тренер заполняет страницу 'План недели'.",
        "2. Клиент после тренировки заполняет 'Журнал тренировок': факт веса, повторов, RPE и комментарий.",
        "3. Один раз в неделю обновляются 'Прогресс тела', 'Фото и состав тела' и 'Самочувствие'.",
        "4. 'Дашборд' показывает динамику веса, талии, выполнения тренировок и восстановления.",
        "5. При боли, RPE 9-10 несколько тренировок подряд или плохом сне не повышать нагрузку.",
    ]
    for idx, text in enumerate(usage, 12):
        ws.cell(idx, 1, text)
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80
    set_common(ws)
    return ws


def setup_profile(wb):
    ws = wb.create_sheet("Профиль клиента")
    ws.merge_cells("A1:F1")
    style_title(ws, "A1", "Профиль клиента")
    profile = [
        ("Имя", "Алексей Иванов", "Цель", "Рекомпозиция: сила + снижение жира без потери мышц"),
        ("Пол", "Мужчина", "Дата старта", START_DATE.strftime("%d.%m.%Y")),
        ("Возраст", 34, "План", "4 силовые + 1 восстановительная зона 2"),
        ("Рост", "184 см", "Текущий вес", "85 кг"),
        ("Опыт", "Опытный спортсмен", "Уровень", "Продвинутый"),
        ("Противопоказания", "Нет", "Ограничения", "Нет заявленных"),
        ("Оборудование", "Полный тренажерный зал", "Предпочтения", "Силовая работа, базовые движения"),
        ("Сон цель", "7.5-8.5 ч", "Шаги цель", "8500-10000 / день"),
        ("Белок цель", "160-180 г / день", "Вода цель", "2.8-3.5 л / день"),
    ]
    for r, row in enumerate(profile, 3):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
            ws.cell(r, c).fill = fill(COLORS["blue"] if c in [2, 4] else COLORS["gray"])
            ws.cell(r, c).font = Font(bold=c in [1, 3])

    ws.merge_cells("A14:F14")
    style_section(ws, "A14", "Коучинговые заметки")
    notes = [
        ["Главный фокус", "Прогрессировать в базовых упражнениях без ухода в хроническую усталость."],
        ["Контроль риска", "Следить за RPE, качеством сна, болью в суставах и падением производительности."],
        ["Правило прогрессии", "Если верх диапазона повторов выполнен при RPE <= 8 и техника стабильна, повышать вес на 2.5-5%."],
        ["Разгрузка", "Каждая 4-я неделя: снизить объем на 30-40%, оставить технику и легкую интенсивность."],
    ]
    for r, row in enumerate(notes, 15):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(r, 2).fill = fill(COLORS["purple"])

    widths = [22, 28, 22, 42, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    set_common(ws)
    return ws


def setup_week_plan(wb):
    ws = wb.create_sheet("План недели")
    ws.merge_cells("A1:Q1")
    style_title(ws, "A1", "Тренировочный план: неделя 1")
    headers = [
        "Дата", "День", "Тренировка", "Блок", "Упражнение", "Мышцы", "Подходы", "Повторы",
        "Вес план", "RPE цель", "RIR", "Темп", "Отдых", "Интенсивность", "Заметки техники",
        "Альтернатива", "Статус",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    plan = [
        (0, "Пн", "Upper Strength", "Разминка", "Гребля + мобилизация плеч", "Общее", 1, "8-10 мин", "-", "5", "4", "контроль", "-", "Низкая", "Подготовить плечи и грудной отдел", "Эллипс", "Запланировано"),
        (0, "Пн", "Upper Strength", "Сила", "Жим штанги лежа", "Грудь/трицепс", 5, "4-6", "90 кг", "8", "2", "3-1-1-0", "2-3 мин", "Высокая", "Лопатки сведены, стопы в пол, без отбива", "Жим гантелей", "Запланировано"),
        (0, "Пн", "Upper Strength", "Сила", "Подтягивания с весом", "Спина/бицепс", 5, "4-6", "+15 кг", "8", "2", "2-1-1-1", "2-3 мин", "Высокая", "Подбородок выше перекладины, без рывков", "Тяга верхнего блока", "Запланировано"),
        (0, "Пн", "Upper Strength", "Гипертрофия", "Жим гантелей на наклонной", "Грудь", 3, "8-10", "32 кг", "8", "2", "3-0-1-0", "90 сек", "Средняя", "Контроль амплитуды, плечи не поднимать", "Хаммер-жим", "Запланировано"),
        (0, "Пн", "Upper Strength", "Гипертрофия", "Тяга штанги в наклоне", "Спина", 4, "6-8", "80 кг", "8", "2", "2-1-1-1", "2 мин", "Средняя", "Корпус стабилен, тянуть локтями", "Тяга T-грифа", "Запланировано"),
        (0, "Пн", "Upper Strength", "Аксессуары", "Разведения в стороны", "Средняя дельта", 3, "12-15", "12 кг", "8", "2", "2-1-2-0", "60 сек", "Средняя", "Не раскачиваться, локоть чуть выше кисти", "Кроссовер", "Запланировано"),
        (1, "Вт", "Lower Strength", "Разминка", "Велотренажер + тазобедренные", "Общее", 1, "10 мин", "-", "5", "4", "контроль", "-", "Низкая", "Подготовить колени, голеностоп, таз", "Дорожка наклон", "Запланировано"),
        (1, "Вт", "Lower Strength", "Сила", "Присед со штангой", "Квадрицепс/ягодицы", 5, "3-5", "130 кг", "8", "2", "3-1-1-0", "3 мин", "Высокая", "Глубина стабильная, колени по линии стоп", "Фронтальный присед", "Запланировано"),
        (1, "Вт", "Lower Strength", "Сила", "Румынская тяга", "Задняя цепь", 4, "6-8", "110 кг", "8", "2", "3-1-1-0", "2-3 мин", "Высокая", "Нейтральная спина, таз назад", "Сгибание ног + гиперэкстензия", "Запланировано"),
        (1, "Вт", "Lower Strength", "Гипертрофия", "Жим ногами", "Ноги", 3, "10-12", "220 кг", "8", "2", "3-0-1-0", "2 мин", "Средняя", "Не отрывать таз, контролировать колени", "Гакк-присед", "Запланировано"),
        (1, "Вт", "Lower Strength", "Аксессуары", "Сгибание ног лежа", "Бицепс бедра", 3, "10-12", "55 кг", "8", "2", "2-1-2-0", "75 сек", "Средняя", "Пауза в сокращении", "Сгибание сидя", "Запланировано"),
        (1, "Вт", "Lower Strength", "Кор", "Планка с весом", "Кор", 3, "45-60 сек", "+15 кг", "7", "3", "статика", "60 сек", "Средняя", "Не проваливать поясницу", "Dead bug", "Запланировано"),
        (2, "Ср", "Recovery", "Кардио", "Зона 2", "Сердечно-сосудистая", 1, "35-45 мин", "Пульс 120-140", "5-6", "4", "ровно", "-", "Низкая", "Дыхание контролируемое, без закисления", "Быстрая ходьба", "Запланировано"),
        (2, "Ср", "Recovery", "Мобилити", "Растяжка бедра/грудного отдела", "Мобилити", 1, "15 мин", "-", "4", "5", "медленно", "-", "Низкая", "Без боли, мягкая амплитуда", "Йога flow", "Запланировано"),
        (3, "Чт", "Push/Pull Hypertrophy", "Гипертрофия", "Жим гантелей сидя", "Плечи", 4, "8-10", "28 кг", "8", "2", "3-0-1-0", "90 сек", "Средняя", "Не прогибаться в пояснице", "Жим в тренажере", "Запланировано"),
        (3, "Чт", "Push/Pull Hypertrophy", "Гипертрофия", "Тяга горизонтального блока", "Спина", 4, "8-12", "75 кг", "8", "2", "2-1-1-1", "90 сек", "Средняя", "Пауза в конце тяги", "Тяга гантели", "Запланировано"),
        (3, "Чт", "Push/Pull Hypertrophy", "Гипертрофия", "Отжимания на брусьях с весом", "Грудь/трицепс", 3, "8-10", "+20 кг", "8", "2", "3-0-1-0", "90 сек", "Средняя", "Не заваливать плечи вперед", "Жим узким хватом", "Запланировано"),
        (3, "Чт", "Push/Pull Hypertrophy", "Аксессуары", "Face pull", "Задняя дельта", 3, "15-20", "25 кг", "7", "3", "2-1-2-1", "60 сек", "Средняя", "Локти высоко, контроль лопаток", "Разведения в наклоне", "Запланировано"),
        (3, "Чт", "Push/Pull Hypertrophy", "Руки", "Суперсет: бицепс + трицепс", "Руки", 3, "10-12", "умеренно", "8", "2", "2-0-2-0", "60 сек", "Средняя", "Без читинга", "Канат/гантели", "Запланировано"),
        (4, "Пт", "Rest", "Восстановление", "Отдых / шаги", "Общее", 1, "8500-10000 шагов", "-", "3", "5", "-", "-", "Низкая", "Сон 8 ч, прогулка, вода", "Легкая мобилити", "Запланировано"),
        (5, "Сб", "Lower Hypertrophy", "Гипертрофия", "Фронтальный присед", "Квадрицепс", 4, "6-8", "95 кг", "8", "2", "3-1-1-0", "2 мин", "Средняя", "Вертикальный корпус, локти высоко", "Гакк-присед", "Запланировано"),
        (5, "Сб", "Lower Hypertrophy", "Гипертрофия", "Болгарские выпады", "Ягодицы/ноги", 3, "8-10/нога", "2x26 кг", "8", "2", "3-0-1-0", "90 сек", "Средняя", "Колено стабильно, корпус слегка вперед", "Выпады назад", "Запланировано"),
        (5, "Сб", "Lower Hypertrophy", "Гипертрофия", "Ягодичный мост", "Ягодицы", 4, "8-12", "140 кг", "8", "2", "2-1-1-1", "2 мин", "Средняя", "Пауза вверху, ребра вниз", "Хип-траст машина", "Запланировано"),
        (5, "Сб", "Lower Hypertrophy", "Аксессуары", "Разгибание ног", "Квадрицепс", 3, "12-15", "65 кг", "8", "2", "2-1-2-0", "75 сек", "Средняя", "Контроль без рывков", "Сисси-присед", "Запланировано"),
        (5, "Сб", "Lower Hypertrophy", "Кор", "Подъем ног в висе", "Пресс", 3, "10-15", "свой вес", "8", "2", "2-1-2-0", "60 сек", "Средняя", "Без раскачки", "Подъем коленей", "Запланировано"),
        (6, "Вс", "Rest", "Восстановление", "Полный отдых", "Общее", 1, "-", "-", "2-3", "5", "-", "-", "Низкая", "Прогулка по желанию, подготовка питания", "Мобилити 10 мин", "Запланировано"),
    ]
    for r, item in enumerate(plan, 4):
        day_offset, *rest = item
        ws.cell(r, 1, START_DATE + timedelta(days=day_offset))
        for c, value in enumerate(rest, 2):
            ws.cell(r, c, value)
        ws.cell(r, 1).number_format = "DD.MM.YYYY"
    add_table(ws, "PlanWeekTable", f"A3:Q{len(plan) + 3}")
    add_dropdown(ws, f"Q4:Q{len(plan) + 3}", ["Запланировано", "Выполнено", "Частично", "Пропущено", "Заменено"])
    add_dropdown(ws, f"N4:N{len(plan) + 3}", ["Низкая", "Средняя", "Высокая"])
    widths = [12, 10, 20, 16, 28, 18, 10, 12, 14, 10, 8, 12, 12, 14, 42, 24, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def setup_training_log(wb):
    ws = wb.create_sheet("Журнал тренировок")
    ws.merge_cells("A1:S1")
    style_title(ws, "A1", "Журнал выполнения тренировок")
    headers = [
        "Дата", "Тренировка", "Упражнение", "План подходы", "План повторы", "План вес",
        "Факт подходы", "Факт повторы", "Факт вес", "Факт RPE", "Боль", "Самочувствие",
        "Техника", "Тоннаж", "% выполнения", "Комментарий клиента", "Комментарий тренера",
        "Решение", "Следующий вес",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    exercises = [
        (START_DATE, "Upper Strength", "Жим штанги лежа", 5, "4-6", "90 кг"),
        (START_DATE, "Upper Strength", "Подтягивания с весом", 5, "4-6", "+15 кг"),
        (START_DATE + timedelta(days=1), "Lower Strength", "Присед со штангой", 5, "3-5", "130 кг"),
        (START_DATE + timedelta(days=1), "Lower Strength", "Румынская тяга", 4, "6-8", "110 кг"),
        (START_DATE + timedelta(days=3), "Push/Pull", "Жим гантелей сидя", 4, "8-10", "28 кг"),
        (START_DATE + timedelta(days=3), "Push/Pull", "Тяга горизонтального блока", 4, "8-12", "75 кг"),
        (START_DATE + timedelta(days=5), "Lower Hypertrophy", "Фронтальный присед", 4, "6-8", "95 кг"),
        (START_DATE + timedelta(days=5), "Lower Hypertrophy", "Ягодичный мост", 4, "8-12", "140 кг"),
    ]
    for r, row in enumerate(exercises, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        ws.cell(r, 14, f'=IF(AND(G{r}<>"",H{r}<>"",I{r}<>""),G{r}*H{r}*VALUE(SUBSTITUTE(SUBSTITUTE(I{r}," кг",""),"+","")),"")')
        ws.cell(r, 15, f'=IF(G{r}<>"",G{r}/D{r},"")')
        ws.cell(r, 15).number_format = "0%"
    add_dropdown(ws, "K4:K60", ["Нет", "Легкий дискомфорт", "Средний дискомфорт", "Сильная боль"])
    add_dropdown(ws, "L4:L60", ["Отличное", "Хорошее", "Среднее", "Плохое"])
    add_dropdown(ws, "M4:M60", ["Отличная", "Хорошая", "Средняя", "Требует правки"])
    add_dropdown(ws, "R4:R60", ["Повысить вес", "Оставить", "Снизить", "Увеличить повторы", "Заменить упражнение", "Проверить технику", "Разгрузка"])
    ws.conditional_formatting.add("J4:J60", CellIsRule(operator="greaterThanOrEqual", formula=["9"], fill=fill(COLORS["red"])))
    ws.conditional_formatting.add("K4:K60", FormulaRule(formula=['OR($K4="Средний дискомфорт",$K4="Сильная боль")'], fill=fill(COLORS["red"])))
    add_table(ws, "TrainingLogTable", "A3:S60")
    widths = [12, 22, 28, 12, 12, 12, 12, 12, 12, 10, 18, 14, 16, 12, 13, 32, 32, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def setup_body_progress(wb):
    ws = wb.create_sheet("Прогресс тела")
    ws.merge_cells("A1:R1")
    style_title(ws, "A1", "Прогресс тела и замеры")
    headers = [
        "Дата", "Неделя", "Вес", "Средний вес 7д", "Талия", "Живот", "Грудь", "Бедра",
        "Ягодицы", "Бедро Л", "Бедро П", "Рука Л", "Рука П", "% жира", "Мышцы кг",
        "Висц. жир", "Комментарий клиента", "Вывод тренера",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    sample = [
        [START_DATE, 1, 85.0, "=AVERAGE(C4:C10)", 84.0, 88.0, 106.0, 98.0, 101.0, 61.0, 61.0, 37.0, 37.0, 16.5, 42.0, 7, "Стартовые замеры", "База для сравнения"],
        [START_DATE + timedelta(days=7), 2, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [START_DATE + timedelta(days=14), 3, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        [START_DATE + timedelta(days=21), 4, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r, row in enumerate(sample, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        ws.cell(r, 1).number_format = "DD.MM.YYYY"
    add_table(ws, "BodyProgressTable", "A3:R20")
    widths = [12, 10, 10, 16, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 12, 34, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def setup_photos(wb):
    ws = wb.create_sheet("Фото и состав тела")
    ws.merge_cells("A1:J1")
    style_title(ws, "A1", "Фото, скриншоты состава тела и визуальная динамика")
    headers = ["Дата", "Неделя", "Фото спереди", "Фото сбоку", "Фото сзади", "Скрин состава тела", "Вес", "Условия", "Комментарий", "Вывод тренера"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    for i in range(4):
        r = 4 + i
        ws.cell(r, 1, START_DATE + timedelta(days=i * 7))
        ws.cell(r, 2, i + 1)
        ws.cell(r, 3, "Вставить ссылку")
        ws.cell(r, 4, "Вставить ссылку")
        ws.cell(r, 5, "Вставить ссылку")
        ws.cell(r, 6, "Вставить ссылку")
        ws.cell(r, 7, "85.0 кг" if i == 0 else "")
        ws.cell(r, 8, "Утро, натощак, одинаковый свет")
        ws.cell(r, 1).number_format = "DD.MM.YYYY"
    ws["A10"] = "Правила фото"
    ws["A10"].font = Font(bold=True, size=12)
    rules = ["Одинаковый свет", "Одинаковая дистанция", "Фото: спереди / сбоку / сзади", "Утром натощак", "Без сильной накачки после тренировки"]
    for i, rule in enumerate(rules, 11):
        ws.cell(i, 1, f"- {rule}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    add_table(ws, "PhotosTable", "A3:J8")
    widths = [12, 10, 22, 22, 22, 26, 10, 32, 32, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def setup_wellbeing(wb):
    ws = wb.create_sheet("Самочувствие")
    ws.merge_cells("A1:N1")
    style_title(ws, "A1", "Самочувствие, восстановление и активность")
    headers = ["Дата", "Сон ч", "Качество сна", "Стресс", "Энергия", "Настроение", "Голод", "DOMS", "Шаги", "Вода л", "Тренировка", "Кардио", "Комментарий", "Флаг риска"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    for i in range(14):
        r = 4 + i
        ws.cell(r, 1, START_DATE + timedelta(days=i))
        ws.cell(r, 14, f'=IF(OR(B{r}<6,D{r}>=8,E{r}<=4,H{r}>=8),"Внимание","ОК")')
        ws.cell(r, 1).number_format = "DD.MM.YYYY"
    add_dropdown(ws, "K4:K80", ["Да", "Нет"])
    add_dropdown(ws, "L4:L80", ["Да", "Нет"])
    ws.conditional_formatting.add("N4:N80", FormulaRule(formula=['$N4="Внимание"'], fill=fill(COLORS["yellow"])))
    add_table(ws, "WellbeingTable", "A3:N80")
    widths = [12, 10, 14, 10, 10, 12, 10, 10, 10, 10, 12, 10, 34, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def setup_dashboard(wb):
    ws = wb.create_sheet("Дашборд")
    ws.merge_cells("A1:H1")
    style_title(ws, "A1", "Дашборд клиента")
    kpis = [
        ("Стартовый вес", "='Прогресс тела'!C4"),
        ("Текущий вес", "=LOOKUP(2,1/('Прогресс тела'!C4:C20<>\"\"),'Прогресс тела'!C4:C20)"),
        ("Изменение веса", "=B4-B3"),
        ("Стартовая талия", "='Прогресс тела'!E4"),
        ("Текущая талия", "=LOOKUP(2,1/('Прогресс тела'!E4:E20<>\"\"),'Прогресс тела'!E4:E20)"),
        ("Изменение талии", "=B7-B6"),
        ("Средний RPE", "=AVERAGEIF('Журнал тренировок'!J4:J60,\">0\")"),
        ("Средний сон", "=AVERAGEIF('Самочувствие'!B4:B80,\">0\")"),
        ("Средние шаги", "=AVERAGEIF('Самочувствие'!I4:I80,\">0\")"),
    ]
    ws["A3"] = "KPI"
    ws["B3"] = "Значение"
    style_header_row(ws, 3, 1, 2)
    for r, (label, formula) in enumerate(kpis, 4):
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        ws.cell(r, 2).fill = fill(COLORS["gray"])
    ws["D3"] = "Последний вывод тренера"
    ws["D3"].font = Font(bold=True, color=COLORS["white"])
    ws["D3"].fill = fill(COLORS["navy"])
    ws.merge_cells("D4:H7")
    ws["D4"] = "Неделя 1: стартовый блок. Контролировать технику, RPE и восстановление. Не форсировать веса при недосыпе."
    ws["D4"].fill = fill(COLORS["purple"])
    ws["D4"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["D9"] = "Фокус следующей недели"
    ws["D9"].font = Font(bold=True, color=COLORS["white"])
    ws["D9"].fill = fill(COLORS["navy"])
    ws.merge_cells("D10:H12")
    ws["D10"] = "Если все базовые упражнения выполнены в верхней границе повторов при RPE <= 8, добавить 2.5-5% веса."
    ws["D10"].fill = fill(COLORS["purple"])
    ws["D10"].alignment = Alignment(wrap_text=True, vertical="top")

    chart = LineChart()
    chart.title = "Вес и талия"
    chart.y_axis.title = "Значение"
    chart.x_axis.title = "Неделя"
    data = Reference(wb["Прогресс тела"], min_col=3, max_col=5, min_row=3, max_row=20)
    cats = Reference(wb["Прогресс тела"], min_col=2, min_row=4, max_row=20)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 15
    ws.add_chart(chart, "A15")

    bar = BarChart()
    bar.title = "Самочувствие: сон и шаги"
    data2 = Reference(wb["Самочувствие"], min_col=2, max_col=9, min_row=3, max_row=17)
    cats2 = Reference(wb["Самочувствие"], min_col=1, min_row=4, max_row=17)
    bar.add_data(data2, titles_from_data=True)
    bar.set_categories(cats2)
    bar.height = 7
    bar.width = 15
    ws.add_chart(bar, "D15")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    set_common(ws)
    return ws


def setup_exercises(wb):
    ws = wb.create_sheet("База упражнений")
    ws.merge_cells("A1:I1")
    style_title(ws, "A1", "База упражнений")
    headers = ["Упражнение", "Категория", "Мышцы", "Оборудование", "Сложность", "Ключ техники", "Ошибка", "Альтернатива", "Ссылка видео"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header_row(ws, 3, 1, len(headers))
    rows = [
        ["Жим штанги лежа", "Грудь", "Грудь, трицепс", "Штанга", "Продвинутый", "Лопатки сведены, стопы в пол", "Отбив от груди", "Жим гантелей", ""],
        ["Присед со штангой", "Ноги", "Квадрицепс, ягодицы", "Штанга", "Продвинутый", "Колени по линии стоп", "Потеря нейтрали спины", "Фронтальный присед", ""],
        ["Румынская тяга", "Задняя цепь", "Бицепс бедра, ягодицы", "Штанга", "Средний", "Таз назад, спина нейтральна", "Округление поясницы", "Гиперэкстензия", ""],
        ["Подтягивания с весом", "Спина", "Широчайшие, бицепс", "Турник", "Продвинутый", "Контроль лопаток", "Рывки корпусом", "Тяга верхнего блока", ""],
        ["Ягодичный мост", "Ягодицы", "Ягодичные", "Штанга", "Средний", "Пауза вверху", "Переразгибание поясницы", "Хип-траст машина", ""],
    ]
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    add_table(ws, "ExerciseBaseTable", "A3:I30")
    widths = [26, 16, 24, 18, 16, 34, 30, 24, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    set_common(ws)
    return ws


def polish_workbook(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    if cell.row > 1 and cell.fill.fill_type is None:
                        cell.fill = fill(COLORS["white"])
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    for ws_name in ["План недели", "Журнал тренировок", "Прогресс тела", "Самочувствие"]:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if cell.column in [7, 8, 9, 10, 11, 12, 13] and ws_name in ["План недели", "Журнал тренировок"]:
                    cell.fill = fill(COLORS["light_blue"])
                if ws_name == "Журнал тренировок" and cell.column in [16, 17, 18, 19]:
                    cell.fill = fill(COLORS["purple"])

    wb.active = wb.sheetnames.index("Инструкция")


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    setup_instruction(wb)
    setup_profile(wb)
    setup_week_plan(wb)
    setup_training_log(wb)
    setup_body_progress(wb)
    setup_photos(wb)
    setup_wellbeing(wb)
    setup_dashboard(wb)
    setup_exercises(wb)
    polish_workbook(wb)
    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    main()
