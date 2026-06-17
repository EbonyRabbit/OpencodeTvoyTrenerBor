from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


FILE = "Клиент_Алексей_Иванов_ведение.xlsx"
thin = Side(style="thin", color="D9D9D9")


wb = load_workbook(FILE)

for name in [n for n in wb.sheetnames if n.startswith("Гип W") or n.startswith("Инт W")]:
    ws = wb[name]

    merged = list(ws.merged_cells.ranges)
    for rng in merged:
        if rng.max_col > 13:
            min_row, max_row = rng.min_row, rng.max_row
            min_col = rng.min_col
            ws.unmerge_cells(str(rng))
            ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=13,
            )

    if ws.max_column > 13:
        ws.delete_cols(14, ws.max_column - 13)

    for row in range(1, ws.max_row + 1):
        for col in range(1, 14):
            cell = ws.cell(row, col)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [17, 31, 11, 12, 12, 10, 11, 10, 11, 12, 11, 10, 30]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

wb.save(FILE)
