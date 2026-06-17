from datetime import date, timedelta

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


FILE = "Клиент_Алексей_Иванов_ведение.xlsx"
START_DATE = date(2026, 6, 1)

COLORS = {
    "navy": "1F4E78",
    "dark_blue": "17365D",
    "blue": "D9EAF7",
    "light_blue": "EAF4FB",
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "red": "F4CCCC",
    "purple": "EADCF8",
    "gray": "E7E6E6",
    "white": "FFFFFF",
    "black": "000000",
}

thin = Side(style="thin", color="D9D9D9")
medium = Side(style="medium", color="9E9E9E")


def fill(color):
    return PatternFill("solid", fgColor=color)


def border(kind="thin"):
    side = medium if kind == "medium" else thin
    return Border(top=side, bottom=side, left=side, right=side)


def style(cell, bg=None, bold=False, color="000000", size=10, center=False):
    cell.font = Font(bold=bold, color=color, size=size)
    if bg:
        cell.fill = fill(bg)
    cell.border = border()
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )


def add_dropdown(ws, cell_range, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def weight(value, multiplier):
    if value is None:
        return "-"
    if isinstance(value, str):
        return value
    rounded = round(value * multiplier / 2.5) * 2.5
    return f"{rounded:g} кг"


def plus_weight(value, multiplier):
    if isinstance(value, str):
        return value
    rounded = round(value * multiplier / 2.5) * 2.5
    return f"+{rounded:g} кг"


def clear_old_training_sheets(wb):
    remove = [
        "План недели",
        "Журнал тренировок",
        "Навигация тренировок",
    ]
    remove += [f"Гип W{i}" for i in range(1, 9)]
    remove += [f"Инт W{i}" for i in range(1, 9)]
    for name in remove:
        if name in wb.sheetnames:
            del wb[name]


def create_nav(wb):
    ws = wb.create_sheet("Навигация тренировок", 2)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Навигация по тренировочным циклам"
    style(ws["A1"], COLORS["navy"], True, COLORS["white"], 16, True)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = "План и журнал теперь совмещены: клиент заполняет факт сразу справа от запланированного упражнения. Не нужно отдельно переходить в журнал."
    style(ws["A2"], COLORS["gray"], False, COLORS["black"], 10, False)
    ws.row_dimensions[2].height = 32

    headers = ["Цикл", "Неделя", "Фокус", "Период", "Открыть", "Что контролировать"]
    for col, header in enumerate(headers, 1):
        c = ws.cell(4, col, header)
        style(c, COLORS["navy"], True, COLORS["white"], 10, True)

    rows = []
    for week in range(1, 9):
        start = START_DATE + timedelta(days=(week - 1) * 7)
        end = start + timedelta(days=6)
        rows.append([
            "Гипертрофия",
            week,
            hypertrophy_focus(week),
            f"{start:%d.%m.%Y} - {end:%d.%m.%Y}",
            f"Гип W{week}",
            "RPE, техника, сон, прогрессия веса/повторов",
        ])
    base = START_DATE + timedelta(days=8 * 7)
    for week in range(1, 9):
        start = base + timedelta(days=(week - 1) * 7)
        end = start + timedelta(days=6)
        rows.append([
            "Выносливость + интервальная силовая",
            week,
            interval_focus(week),
            f"{start:%d.%m.%Y} - {end:%d.%m.%Y}",
            f"Инт W{week}",
            "Пульс, плотность работы, техника на усталости",
        ])

    for r, row in enumerate(rows, 5):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value)
            bg = COLORS["light_blue"] if row[0] == "Гипертрофия" else COLORS["green"]
            style(cell, bg if c <= 4 else COLORS["white"], False, COLORS["black"], 10, c in [2])
        link_cell = ws.cell(r, 5)
        link_cell.hyperlink = f"#'{row[4]}'!A1"
        link_cell.style = "Hyperlink"
        link_cell.value = "Перейти"
        style(link_cell, COLORS["yellow"], True, COLORS["black"], 10, True)

    ws.merge_cells("A23:H23")
    ws["A23"] = "Как добавлять будущие недели"
    style(ws["A23"], COLORS["navy"], True, COLORS["white"], 12, False)
    tips = [
        "1. Дублируй последний недельный лист, чтобы сохранить формат и выпадающие списки.",
        "2. Меняй дату, фокус недели и плановые веса/повторы по логике прогрессии.",
        "3. Старые недели не удаляй: это история прогресса клиента.",
        "4. На навигации добавляй строку со ссылкой на новую неделю, чтобы не листать файл вручную.",
    ]
    for i, text in enumerate(tips, 24):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws.cell(i, 1, text)
        style(ws.cell(i, 1), COLORS["gray"], False, COLORS["black"], 10, False)

    widths = [30, 10, 34, 23, 14, 44, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def hypertrophy_focus(week):
    focuses = {
        1: "Адаптация к объему, техника, RPE 7",
        2: "+1 повтор или +2.5% веса, RPE 7-8",
        3: "Увеличение объема, тяжелые базовые RPE 8",
        4: "Разгрузка: -35% объема, RPE 6",
        5: "Новый блок: тяжелее, диапазон 6-10",
        6: "Рост объема: +1 подход в ключевых движениях",
        7: "Пиковая гипертрофия, RPE 8-9 без отказа",
        8: "Консолидация и контрольные AMRAP без отказа",
    }
    return focuses[week]


def interval_focus(week):
    focuses = {
        1: "База выносливости, интервалы умеренно",
        2: "+1 раунд в основных комплексах",
        3: "Сокращение отдыха, выше плотность",
        4: "Разгрузка: меньше раундов, техника",
        5: "Новый блок: тяжелее интервальная сила",
        6: "Пороговая работа и плотность",
        7: "Пиковая неделя: максимум качества",
        8: "Тестовая неделя: контрольные комплексы",
    }
    return focuses[week]


def day_header(ws, row, title, goal):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    ws.cell(row, 1, title)
    style(ws.cell(row, 1), COLORS["dark_blue"], True, COLORS["white"], 12, False)
    ws.row_dimensions[row].height = 26

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=16)
    ws.cell(row + 1, 1, goal)
    style(ws.cell(row + 1, 1), COLORS["yellow"], True, COLORS["black"], 10, False)
    ws.row_dimensions[row + 1].height = 28

    headers = [
        "Раздел", "Упражнение", "План подходы", "План повторы", "План вес", "RPE/RIR", "Темп", "Отдых",
        "Факт подходы", "Факт повторы", "Факт вес", "Факт RPE", "Боль", "Комментарий клиента", "Решение тренера", "След. шаг",
    ]
    for col, header in enumerate(headers, 1):
        c = ws.cell(row + 2, col, header)
        bg = COLORS["navy"] if col <= 8 else COLORS["dark_blue"]
        style(c, bg, True, COLORS["white"], 9, True)
    ws.row_dimensions[row + 2].height = 34
    return row + 3


def add_exercise_rows(ws, row, exercises):
    for data in exercises:
        for col, value in enumerate(data, 1):
            c = ws.cell(row, col, value)
            if col <= 8:
                bg = COLORS["light_blue"]
            elif col in [9, 10, 11, 12, 13, 14]:
                bg = COLORS["blue"]
            else:
                bg = COLORS["purple"]
            style(c, bg, False, COLORS["black"], 9, col in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])
        ws.row_dimensions[row].height = 36
        row += 1
    return row + 1


def create_week_sheet(wb, name, title, subtitle, days, index):
    ws = wb.create_sheet(name, index)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:P1")
    ws["A1"] = title
    style(ws["A1"], COLORS["navy"], True, COLORS["white"], 15, True)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:P2")
    ws["A2"] = subtitle
    style(ws["A2"], COLORS["gray"], False, COLORS["black"], 10, False)
    ws.row_dimensions[2].height = 28

    ws["A3"] = "Назад к навигации"
    ws["A3"].hyperlink = "#'Навигация тренировок'!A1"
    ws["A3"].style = "Hyperlink"
    style(ws["A3"], COLORS["yellow"], True, COLORS["black"], 10, True)
    ws.merge_cells("B3:P3")
    ws["B3"] = "Заполнять после каждого упражнения: факт подходов, повторов, веса, RPE, боль и короткий комментарий."
    style(ws["B3"], COLORS["yellow"], False, COLORS["black"], 10, False)

    row = 5
    for day_title, goal, exercises in days:
        start = day_header(ws, row, day_title, goal)
        row = add_exercise_rows(ws, start, exercises)

    widths = [17, 31, 11, 12, 12, 10, 11, 10, 11, 12, 11, 10, 16, 28, 22, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    add_dropdown(ws, f"M5:M{row + 20}", ["Нет", "Легкий дискомфорт", "Средний дискомфорт", "Сильная боль"])
    add_dropdown(ws, f"O5:O{row + 20}", ["Повысить вес", "Оставить", "Снизить", "Добавить повтор", "Сократить отдых", "Заменить", "Разгрузка"])
    ws.conditional_formatting.add(f"L5:L{row + 20}", CellIsRule(operator="greaterThanOrEqual", formula=["9"], fill=fill(COLORS["red"])))
    ws.conditional_formatting.add(f"M5:M{row + 20}", FormulaRule(formula=['OR($M5="Средний дискомфорт",$M5="Сильная боль")'], fill=fill(COLORS["red"])))

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def hypertrophy_days(week):
    multipliers = {1: 1.00, 2: 1.025, 3: 1.05, 4: 0.90, 5: 1.075, 6: 1.10, 7: 1.125, 8: 1.05}
    set_adj = {1: 0, 2: 0, 3: 1, 4: -1, 5: 0, 6: 1, 7: 1, 8: 0}
    reps_main = {1: "8-10", 2: "9-11", 3: "8-12", 4: "8-10", 5: "6-9", 6: "7-10", 7: "8-10", 8: "AMRAP 8-12"}
    rpe = {1: "7/3", 2: "7-8/2-3", 3: "8/2", 4: "6/4", 5: "8/2", 6: "8/2", 7: "8-9/1-2", 8: "8/2"}
    m = multipliers[week]
    add = set_adj[week]
    start = START_DATE + timedelta(days=(week - 1) * 7)

    def sets(base):
        return max(2, base + add)

    return [
        (
            f"Понедельник {start:%d.%m.%Y} | Upper A | Грудь + спина",
            f"Неделя {week}: {hypertrophy_focus(week)}.",
            [
                ["Разминка", "Гребля + плечи", 1, "8-10 мин", "-", "5/4", "контроль", "-", "", "", "", "", "", "", "", ""],
                ["База", "Жим штанги лежа", sets(4), reps_main[week], weight(90, m), rpe[week], "3-1-1-0", "2-3 мин", "", "", "", "", "", "", "", ""],
                ["База", "Подтягивания с весом", sets(4), reps_main[week], plus_weight(15, m), rpe[week], "2-1-1-1", "2-3 мин", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Жим гантелей на наклонной", sets(3), "10-12", weight(32, m), rpe[week], "3-0-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Тяга штанги в наклоне", sets(3), "8-10", weight(80, m), rpe[week], "2-1-1-1", "2 мин", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Разведения в стороны", sets(3), "12-15", weight(12, m), "8/2", "2-1-2-0", "60 сек", "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Вторник {(start + timedelta(days=1)):%d.%m.%Y} | Lower A | Ноги, силовой акцент",
            "Главный контроль: техника приседа и отсутствие боли в коленях/пояснице.",
            [
                ["Разминка", "Велотренажер + тазобедренные", 1, "10 мин", "-", "5/4", "контроль", "-", "", "", "", "", "", "", "", ""],
                ["База", "Присед со штангой", sets(4), "6-8" if week < 5 else "5-8", weight(130, m), rpe[week], "3-1-1-0", "3 мин", "", "", "", "", "", "", "", ""],
                ["База", "Румынская тяга", sets(4), reps_main[week], weight(110, m), rpe[week], "3-1-1-0", "2-3 мин", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Жим ногами", sets(3), "10-15", weight(220, m), "8/2", "3-0-1-0", "2 мин", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Сгибание ног лежа", sets(3), "10-15", weight(55, m), "8/2", "2-1-2-0", "75 сек", "", "", "", "", "", "", "", ""],
                ["Кор", "Планка с весом", 3, "45-60 сек", plus_weight(15, m), "7/3", "статика", "60 сек", "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Четверг {(start + timedelta(days=3)):%d.%m.%Y} | Upper B | Плечи + спина + руки",
            "Объемная работа. Цель: накопить качественные повторы без технического отказа.",
            [
                ["База", "Жим гантелей сидя", sets(4), "8-10", weight(28, m), rpe[week], "3-0-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Тяга горизонтального блока", sets(4), "8-12", weight(75, m), rpe[week], "2-1-1-1", "90 сек", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Брусья с весом", sets(3), "8-12", plus_weight(20, m), "8/2", "3-0-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Face Pull", sets(3), "15-20", weight(25, m), "7/3", "2-1-2-1", "60 сек", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Бицепс + трицепс", sets(3), "10-12", "умеренно", "8/2", "2-0-2-0", "60 сек", "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Суббота {(start + timedelta(days=5)):%d.%m.%Y} | Lower B | Ноги, объем",
            "Объемная работа ног и ягодиц. Не уходить в отказ на выпадах и мосте.",
            [
                ["База", "Фронтальный присед", sets(4), "6-10", weight(95, m), rpe[week], "3-1-1-0", "2 мин", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Болгарские выпады", sets(3), "8-10/нога", weight(26, m), "8/2", "3-0-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Гипертрофия", "Ягодичный мост", sets(4), "8-12", weight(140, m), "8/2", "2-1-1-1", "2 мин", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Разгибание ног", sets(3), "12-15", weight(65, m), "8/2", "2-1-2-0", "75 сек", "", "", "", "", "", "", "", ""],
                ["Кор", "Подъем ног в висе", 3, "10-15", "свой вес", "8/2", "2-1-2-0", "60 сек", "", "", "", "", "", "", "", ""],
            ],
        ),
    ]


def interval_days(week):
    rounds = {1: 4, 2: 5, 3: 5, 4: 3, 5: 5, 6: 6, 7: 6, 8: 4}
    rest = {1: "90 сек", 2: "75 сек", 3: "60 сек", 4: "90 сек", 5: "75 сек", 6: "60 сек", 7: "45-60 сек", 8: "по тесту"}
    intensity = {1: "7/3", 2: "7-8/2-3", 3: "8/2", 4: "6/4", 5: "8/2", 6: "8/2", 7: "8-9/1-2", 8: "8/2"}
    m = {1: .75, 2: .775, 3: .80, 4: .70, 5: .825, 6: .85, 7: .875, 8: .80}[week]
    start = START_DATE + timedelta(days=(8 + week - 1) * 7)
    r = rounds[week]

    return [
        (
            f"Понедельник {start:%d.%m.%Y} | Interval Strength A | Ноги + толкание",
            f"Неделя {week}: {interval_focus(week)}. Работать быстро, но без развала техники.",
            [
                ["Разминка", "Дорожка наклон + мобилити", 1, "10 мин", "-", "5/4", "ровно", "-", "", "", "", "", "", "", "", ""],
                ["Сила", "Фронтальный присед", 4, "5", weight(95, m), intensity[week], "2-0-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Интервал", "Комплекс: Assault Bike 12 ккал + KB Swing 15 + Push-up 15", r, "на время", "24 кг KB", intensity[week], "ритм", rest[week], "", "", "", "", "", "", "", ""],
                ["Кор", "Farmer Walk", 4, "40 м", "2x32 кг", "7/3", "ровно", "60 сек", "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Среда {(start + timedelta(days=2)):%d.%m.%Y} | Aerobic + Core | Зона 2",
            "Цель: развить базовую выносливость и ускорить восстановление между силовыми интервалами.",
            [
                ["Кардио", "Зона 2: велосипед/гребля/дорожка", 1, f"{35 + min(week, 5) * 5} мин", "Пульс 120-145", "5-6/4", "ровно", "-", "", "", "", "", "", "", "", ""],
                ["Кор", "Dead Bug + Side Plank", 3, "12/стор + 40 сек", "свой вес", "6/4", "контроль", "45 сек", "", "", "", "", "", "", "", ""],
                ["Мобилити", "Тазобедренные + грудной отдел", 1, "12-15 мин", "-", "4/5", "медленно", "-", "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Пятница {(start + timedelta(days=4)):%d.%m.%Y} | Interval Strength B | Тяга + переносы",
            "Цель: силовая выносливость задней цепи и спины, контроль дыхания под нагрузкой.",
            [
                ["Сила", "Румынская тяга", 4, "6", weight(110, m), intensity[week], "2-1-1-0", "90 сек", "", "", "", "", "", "", "", ""],
                ["Интервал", "EMOM: гребля 12 ккал / подтягивания 8 / wall ball 15", r * 3, "минут", "по готовности", intensity[week], "ритм", "EMOM", "", "", "", "", "", "", "", ""],
                ["Доп. упражнения", "Тяга саней / sled push", 5, "20 м", "тяжело", "8/2", "мощно", rest[week], "", "", "", "", "", "", "", ""],
            ],
        ),
        (
            f"Суббота {(start + timedelta(days=5)):%d.%m.%Y} | Mixed Conditioning | Интервальный круг",
            "Цель: плотность работы. Сравнивать общий результат недели к неделе, но технику не ломать.",
            [
                ["Круг", "12 ккал Bike + 10 DB Thruster + 12 TRX Row + 200 м run", r, "раундов", "2x16-20 кг", intensity[week], "ровно", rest[week], "", "", "", "", "", "", "", ""],
                ["Финишер", "Battle Rope", 6 + week, "20 сек работа / 40 сек отдых", "-", "8/2", "быстро", "40 сек", "", "", "", "", "", "", "", ""],
                ["Заминка", "Ходьба + дыхание", 1, "8-10 мин", "-", "3/5", "медленно", "-", "", "", "", "", "", "", "", ""],
            ],
        ),
    ]


def update_instruction(wb):
    if "Инструкция" not in wb.sheetnames:
        return
    ws = wb["Инструкция"]
    row = ws.max_row + 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1, "Обновление структуры тренировок")
    style(ws.cell(row, 1), COLORS["navy"], True, COLORS["white"], 12, False)
    notes = [
        "Листы 'План недели' и 'Журнал тренировок' заменены на недельные листы формата план + факт.",
        "Клиент заполняет факт выполнения сразу после каждого упражнения в правой части таблицы.",
        "Для быстрого доступа используй лист 'Навигация тренировок'.",
    ]
    for i, note in enumerate(notes, row + 1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws.cell(i, 1, note)
        style(ws.cell(i, 1), COLORS["gray"], False, COLORS["black"], 10, False)


def main():
    wb = load_workbook(FILE)
    clear_old_training_sheets(wb)
    create_nav(wb)

    index = wb.sheetnames.index("Навигация тренировок") + 1
    for week in range(1, 9):
        start = START_DATE + timedelta(days=(week - 1) * 7)
        create_week_sheet(
            wb,
            f"Гип W{week}",
            f"Гипертрофия | Неделя {week}",
            f"Период: {start:%d.%m.%Y} - {(start + timedelta(days=6)):%d.%m.%Y} | {hypertrophy_focus(week)}",
            hypertrophy_days(week),
            index,
        )
        index += 1

    for week in range(1, 9):
        start = START_DATE + timedelta(days=(8 + week - 1) * 7)
        create_week_sheet(
            wb,
            f"Инт W{week}",
            f"Выносливость + интервальная силовая работа | Неделя {week}",
            f"Период: {start:%d.%m.%Y} - {(start + timedelta(days=6)):%d.%m.%Y} | {interval_focus(week)}",
            interval_days(week),
            index,
        )
        index += 1

    update_instruction(wb)
    wb.active = wb.sheetnames.index("Навигация тренировок")
    wb.save(FILE)


if __name__ == "__main__":
    main()
