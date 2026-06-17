import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CYCLE_NAME = "Подросток 12 лет"
TOTAL_WEEKS = 12
DAYS = ["Вторник", "Четверг", "Суббота"]
DAY_TYPES = ["A — Фулл-боди", "B — Фулл-боди", "C — Фулл-боди"]

PLAN_HEADERS = ["Блок", "Упражнение", "Подходы", "Повторы", "Вес", "Темп", "Отдых", "RPE", "Заметки"]
COL_WIDTHS = [18, 48, 10, 12, 10, 14, 10, 10, 36]
ALL_HEADERS_COUNT = len(PLAN_HEADERS)

NAVY = "1F3864"
NAVY_DARK = "2E5090"
LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
DAY_COLORS = {
    "A": "DAEEF3",
    "B": "E2EFDA",
    "C": "FCE4D6",
}

thin = Side(style="thin", color="999999")

# Phase presets
PHASES = {
    1: {"sets": 2, "reps": "12–15", "weight": "1–2 кг", "tempo": "2-0-1-0", "rest": "60 с", "rpe": "7", "label": "Фаза 1 — Техника"},
    2: {"sets": 3, "reps": "12–15", "weight": "2–3 кг", "tempo": "2-0-1-0", "rest": "45–60 с", "rpe": "7–8", "label": "Фаза 2 — Развитие"},
    3: {"sets": 3, "reps": "10–12", "weight": "2–4.5 кг", "tempo": "3-0-1-0", "rest": "45–60 с", "rpe": "7–8", "label": "Фаза 3 — Укрепление"},
}

PHASE_MAP = {1: 1, 2: 1, 3: 1, 4: 1, 5: 2, 6: 2, 7: 2, 8: 2, 9: 3, 10: 3, 11: 3, 12: 3}

WEEKLY_FOCUS = {
    1: "Освоение техники, лёгкий вес",
    2: "Закрепление формы",
    3: "Уверенное выполнение",
    4: "Контроль + стабильность",
    5: "Увеличение объёма: +1 подход",
    6: "Рост нагрузки",
    7: "Удержание качества",
    8: "Повышение интенсивности",
    9: "Укрепление + контроль",
    10: "Пик формы",
    11: "Стабильность под нагрузкой",
    12: "Завершение цикла",
}

base_exercises = {
    "A": [
        ["Ноги", "Гоблет-приседания с гантелью", "sets", "reps", "weight", "tempo", "rest", "rpe", "Колени по стопам, спина прямая"],
        ["Ноги", "Выпады в стороны с гантелью", "sets", "reps", "weight", "tempo", "rest", "rpe", "Носок наружу, колено по стопе"],
        ["Грудь+Трицепс", "Отжимания узким хватом (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Локти вдоль корпуса"],
        ["Спина+Бицепс", "Тяга гантели одной рукой", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Спина прямо, локоть назад"],
        ["Плечи", "Жим гантелей стоя", "sets", "reps", "weight", "tempo", "rest", "rpe", "Не прогибаться в пояснице"],
        ["Кор", "Планка Копенгагена", "sets", "15–20 с на сторону", "вес тела", "—", "30 с", "rpe", "Верхняя нога на возвышении"],
    ],
    "B": [
        ["Ноги", "Выпады назад с гантелями", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Колено передней ноги за носком"],
        ["Ноги", "Сиси-приседания (у стены/с опорой)", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени назад, пятки можно поднять"],
        ["Грудь+Плечи", "Отжимания от пола (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Корпус прямая линия"],
        ["Спина+Бицепс", "Тяга гантели к поясу (в упоре на колено)", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Локоть выше корпуса"],
        ["Задняя поверхность", "Нордические наклоны (с резинкой/руками)", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени прижаты, спина прямо"],
        ["Кор", "Ротации с резинкой (дровосек)", "sets×2", "reps", "резинка", "—", "rest", "rpe", "Скручивание корпусом, руки прямые"],
    ],
    "C": [
        ["Ягодицы+Ноги", "Ягодичный мост с гантелью на бёдрах", "sets", "reps", "weight", "tempo", "rest", "rpe", "Сокращение 1 сек вверху"],
        ["Ноги", "Обратные нордические", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени на полу, корпус назад"],
        ["Грудь+Трицепс", "Отжимания узким хватом (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Локти вдоль корпуса"],
        ["Плечи", "Боковые подъёмы гантелей", "sets", "reps", "weight", "tempo", "rest", "rpe", "Мизинец вверх, без рывков"],
        ["Спина", "Тяга резинки к лицу (Face Pull)", "sets", "reps", "резинка", "—", "rest", "rpe", "Локти в стороны, лопатки вместе"],
        ["Кор", "Подъёмы ног лёжа (низ живота)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Поясница прижата к полу"],
    ],
}

warmup = [
    ["Круговые движения плечами", "30 с × 2"],
    ["Махи ногами вперёд/назад", "15 × 2"],
    ["Наклоны в стороны", "15 × 2"],
    ["Глубокие приседания без веса", "15"],
    ["Лёгкий бег на месте", "45 с"],
]

cooldown = [
    ["Растяжка квадрицепса стоя", "30 с × 2"],
    ["Наклон к ногам сидя", "45 с"],
    ["Кобра / лёгкий прогиб", "30 с"],
    ["Скручивание лёжа (поза ребёнка)", "45 с"],
]

progression = [
    ["1–4", "12–15", "2", "1–2 кг", "7"],
    ["5–8", "12–15", "3", "2–3 кг", "7–8"],
    ["9–12", "10–12", "3", "2–4.5 кг", "7–8"],
]

def fmt(row):
    return Font(size=9)

def header_font(bold=True, size=9, color=WHITE):
    return Font(bold=bold, size=size, color=color)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def style_cell(cell, font_obj=None, fill_obj=None, alignment_obj=None):
    if font_obj:
        cell.font = font_obj
    if fill_obj:
        cell.fill = fill_obj
    if alignment_obj:
        cell.alignment = alignment_obj
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def write_title(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(bold=True, size=14, color=WHITE), fill(NAVY),
               Alignment(horizontal="left", vertical="center"))

def write_subtitle(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(size=10), fill(YELLOW),
               Alignment(horizontal="left", vertical="center"))
    cell.font = Font(size=10)

def write_day_header(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(bold=True, size=11, color=WHITE), fill(NAVY_DARK),
               Alignment(horizontal="left", vertical="center"))

def write_table_header(ws, row, headers, fill_color=NAVY):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        style_cell(cell, Font(bold=True, size=9, color=WHITE), fill(fill_color),
                   Alignment(horizontal="center", vertical="center", wrap_text=True))

def write_data_row(ws, row, values, color=LIGHT_BLUE):
    for c in range(1, ALL_HEADERS_COUNT + 1):
        cell = ws.cell(row, c, values[c - 1] if c <= len(values) and values[c - 1] is not None and values[c - 1] != "" else "")
        style_cell(cell, Font(size=9), fill(color),
                   Alignment(vertical="center", wrap_text=True))

def write_simple_table(ws, start_row, headers, rows, header_color=NAVY, row_color=LIGHT_BLUE):
    write_table_header(ws, start_row, headers, header_color)
    for i, row_data in enumerate(rows):
        write_data_row(ws, start_row + 1 + i, row_data, row_color)
    return start_row + 1 + len(rows)

def fill_workout(day_key, phase_num):
    p = PHASES[phase_num]
    base = base_exercises[day_key]
    out = []
    for ex in base:
        block, name, sets_k, reps_k, weight_k, tempo_k, rest_k, rpe_k, note = ex
        s = str(p["sets"]) if sets_k == "sets" else f'{p["sets"]}×2'
        rep = str(p["reps"])
        w = str(p["weight"]) if weight_k in ("weight",) else weight_k
        t = str(p["tempo"]) if tempo_k == "tempo" else tempo_k
        r = str(p["rest"]) if rest_k == "rest" else rest_k
        rpe = str(p["rpe"]) if rpe_k == "rpe" else rpe_k
        out.append([block, name, s, rep, w, t, r, rpe, note])
    return out


wb = openpyxl.Workbook()

for week_num in range(1, TOTAL_WEEKS + 1):
    ws = wb.create_sheet(f"W{week_num}")
    total_cols = ALL_HEADERS_COUNT
    phase = PHASE_MAP[week_num]
    phase_name = PHASES[phase]["label"]

    write_title(ws, 1, f"{CYCLE_NAME} | Неделя {week_num} — {phase_name}", total_cols)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    cell = ws.cell(2, 1, f"Фокус: {WEEKLY_FOCUS[week_num]}")
    style_cell(cell, Font(size=10), fill(YELLOW),
               Alignment(horizontal="left", vertical="center"))

    r = 4

    for day_idx, day_name in enumerate(DAYS):
        day_key = ["A", "B", "C"][day_idx]
        day_label = f"{day_name} | {DAY_TYPES[day_idx]}"
        day_color = DAY_COLORS[day_key]

        write_day_header(ws, r, day_label, total_cols)
        r += 1

        write_table_header(ws, r, PLAN_HEADERS)
        r += 1

        exercises = fill_workout(day_key, phase)
        for ex in exercises:
            write_data_row(ws, r, ex, day_color)
            r += 1

        r += 1

    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c - 1]

    ws.print_options.fitToWidth = 1
    ws.page_setup.fitToWidth = 1

# --- Sheet: Инфо ---
ws_info = wb.create_sheet("Инфо", 0)
total_cols = 4
for c in range(1, 5):
    ws_info.column_dimensions[get_column_letter(c)].width = [28, 50, 16, 24][c - 1]

write_title(ws_info, 1, f"{CYCLE_NAME} | 3 × Фулл-боди в неделю", total_cols)

r = 3
ws_info.cell(r, 1, "Инвентарь").font = Font(bold=True, size=11)
r += 1
items = [
    "Гантели разборные (диски 1кг × 4 + 1.25кг × 4)",
    "Резинки разной жёсткости",
    "Коврик",
    "Стул / диван (для опоры)",
]
for item in items:
    ws_info.cell(r, 1, f"• {item}")
    ws_info.cell(r, 1).font = Font(size=10)
    r += 1

r += 2
ws_info.cell(r, 1, "Безопасность").font = Font(bold=True, size=11)
r += 1
safety = [
    "❌ Штанга на спину/плечи (приседания, выпады со штангой)",
    "❌ Жим штанги лёжа",
    "❌ Становая тяга с пола",
    "❌ Любые упражнения с грифом на шее/позвоночнике",
    "❌ Прыжки с отягощением",
    "✅ Все упражнения с гантелями и резинками — безопасны",
    "✅ Вес только при полном контроле техники",
]
for item in safety:
    ws_info.cell(r, 1, item)
    ws_info.cell(r, 1).font = Font(size=10)
    r += 1

r += 2
ws_info.cell(r, 1, "Структура недели").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["День", "Тип тренировки"],
                   [["Пн", "A — Фулл-боди"],
                    ["Ср", "B — Фулл-боди"],
                    ["Пт", "C — Фулл-боди"]],
                   NAVY, LIGHT_BLUE)

r += 7
ws_info.cell(r, 1, "Прогрессия по фазам").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Недели", "Повторы", "Подходы", "Вес", "RPE"],
                   progression, NAVY_DARK, DAY_COLORS["A"])

r += 8
ws_info.cell(r, 1, "Разминка (перед каждой тренировкой)").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Упражнение", "Длительность"],
                   warmup, NAVY, LIGHT_BLUE)

r += 6
ws_info.cell(r, 1, "Заминка").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Упражнение", "Длительность"],
                   cooldown, NAVY_DARK, LIGHT_BLUE)

r += 6
ws_info.cell(r, 1, "Рекомендации для подростка").font = Font(bold=True, size=11)
r += 1
tips = [
    "• Пить воду до/во время/после тренировки",
    "• Спать 8-10 часов",
    "• Есть достаточно белка (рыба, мясо, яйца, творог)",
    "• Если боль — остановиться, сказать взрослому",
    "• Не соревноваться с весом — главное техника",
    "• Между тренировками — минимум 1 день отдыха",
]
for tip in tips:
    ws_info.cell(r, 1, tip)
    ws_info.cell(r, 1).font = Font(size=10)
    r += 1

wb.remove(wb["Sheet"])

filename = "/Users/iuriisho/OpenCode/Подросток_12_недель.xlsx"
wb.save(filename)
print(f"Сохранено: {filename}")
