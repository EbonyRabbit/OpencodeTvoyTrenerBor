from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Strength_Hypertrophy_12_weeks.xlsx"

NAVY = "1F4E78"
WHITE = "FFFFFF"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF4FB"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
ORANGE = "FCE5CD"
RED = "F4CCCC"
LIGHT_GRAY = "E7E6E6"

thin = Side(style="thin", color="D9D9D9")


def fill(color):
    return PatternFill("solid", fgColor=color)


def style_title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=14, color=WHITE)
    ws[cell].fill = fill(NAVY)
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, start, end):
    for col in range(start, end + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def set_common(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                if cell.row > 1 and isinstance(cell.fill, PatternFill) and cell.fill.fgColor.rgb == "00000000":
                    cell.fill = fill(WHITE)


HEADERS = ["Блок", "Упражнение", "Подходы", "Повторы", "Вес/% 1ПМ", "RPE", "Отдых", "Заметки"]

# ============================================================
# PHASE PRESETS
# ============================================================
PHASES = {
    1: {  # Hypertrophy (weeks 1-4)
        "name": "Гипертрофия",
        "main_sets": 4,
        "main_reps": "8-12",
        "main_weight": "70-75%",
        "main_rpe": "7-8",
        "main_rest": "90 с",
        "acc_sets": 3,
        "acc_reps": "12-15",
        "acc_rpe": "7-8",
        "acc_rest": "60 с",
        "arm_sets": 3,
        "arm_reps": "12-15",
        "arm_rpe": "7-8",
        "arm_rest": "45-60 с",
    },
    2: {  # Strength (weeks 5-8)
        "name": "Сила",
        "main_sets": 5,
        "main_reps": "4-6",
        "main_weight": "80-85%",
        "main_rpe": "8-9",
        "main_rest": "180 с",
        "acc_sets": 3,
        "acc_reps": "8-10",
        "acc_rpe": "8",
        "acc_rest": "90 с",
        "arm_sets": 3,
        "arm_reps": "8-10",
        "arm_rpe": "8",
        "arm_rest": "60-90 с",
    },
    3: {  # Mixed (weeks 9-12)
        "name": "Сила + Гипертрофия",
        "main_sets": 4,
        "main_reps": "6-8",
        "main_weight": "75-82%",
        "main_rpe": "8-9",
        "main_rest": "120 с",
        "acc_sets": 3,
        "acc_reps": "10-12",
        "acc_rpe": "8",
        "acc_rest": "60-75 с",
        "arm_sets": 3,
        "arm_reps": "10-12",
        "arm_rpe": "8",
        "arm_rest": "60 с",
    },
}

# ============================================================
# DAY TEMPLATES
# ============================================================
# Each day is a list of (block, exercise, sets, reps, weight, rpe, rest, notes)
# where sets/reps/weight/rpe/rest use phase keys like "main_sets", "main_reps", etc.

def phase_val(p, key, override=None):
    return override if override is not None else p[key]


def fill_day(phase, exercises):
    p = PHASES[phase]
    out = [("Разминка", "Суставная разминка + активация", "", "", "", "", "10 мин", "")]
    for block, ex_name, sets_k, reps_k, weight_k, rpe_k, rest_k, notes in exercises:
        if block == "Сила (основное)":
            block_label = "Сила (основное)"
            s = phase_val(p, sets_k)
            r = phase_val(p, reps_k)
            w = phase_val(p, weight_k) if isinstance(weight_k, str) else weight_k
            rpe = phase_val(p, rpe_k)
            rest = phase_val(p, rest_k)
        elif block == "Вспомогательные":
            block_label = "Вспомогательные"
            s = phase_val(p, sets_k)
            r = phase_val(p, reps_k)
            w = weight_k if weight_k else "рабочий"
            rpe = phase_val(p, rpe_k)
            rest = phase_val(p, rest_k)
        elif block == "Руки":
            block_label = "Руки"
            s = phase_val(p, sets_k)
            r = phase_val(p, reps_k)
            w = weight_k if weight_k else "рабочий"
            rpe = phase_val(p, rpe_k)
            rest = phase_val(p, rest_k)
        else:
            block_label = block
            s = sets_k
            r = reps_k
            w = weight_k
            rpe = rpe_k
            rest = rest_k

        out.append((block_label, ex_name, str(s), str(r), w, rpe, rest, notes))
    out.append(("Заминка", "Растяжка целевых групп", "", "", "", "", "5-10 мин", ""))
    return out


# ============================================================
# DAY 1: UPPER A — горизонтальный жим + тяга + вертикальный баланс
# ============================================================
# Push: 2 (1H + 1V) | Pull: 2 (1H + 1V) | Rear delt: 1
def day1_upper(phase):
    return fill_day(phase, [
        ("Сила (основное)", "Жим гантелей на наклонной скамье", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Горизонтальный жим, угол 30°"),
        ("Сила (основное)", "Тяга горизонтальная одной рукой", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Горизонтальная тяга, корпус стабилен"),
        ("Сила (основное)", "Отжимания на брусьях с весом", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Вертикальный жим, локти вдоль корпуса"),
        ("Сила (основное)", "Вертикальная тяга одной рукой", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Вертикальная тяга, локоть вниз-назад"),
        ("Вспомогательные", "Отведение плеч в сторону (гантели)", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Средняя дельта, мизинец вверх"),
        ("Вспомогательные", "Разведение рук назад (обратная бабочка)", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Задняя дельта, горизонтальное отведение"),
    ])


# ============================================================
# DAY 2: LEGS + ARMS
# ============================================================
def day2_legs_arms(phase):
    return fill_day(phase, [
        ("Сила (основное)", "Приседания с трэп грифом", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Полная амплитуда, колени по траектории стоп"),
        ("Сила (основное)", "Выпады шагающие (гантели)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Шаг средней длины, вес в передней пятке"),
        ("Сила (основное)", "Ягодичный мостик (штанга)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Сокращение 1-2 с в верхней точке"),
        ("Вспомогательные", "Сгибание ног лёжа", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Бицепс бедра, медленная негатива"),
        ("Руки", "Французский жим лёжа (EZ)", "arm_sets", "arm_reps", "", "arm_rpe", "arm_rest", "Локти стабильны, трицепс"),
        ("Руки", "Молотковые сгибания (гантели)", "arm_sets", "arm_reps", "", "arm_rpe", "arm_rest", "Брахиалис + бицепс, нейтральный хват"),
    ])


# ============================================================
# DAY 3: UPPER B — вертикальная тяга + жим + горизонтальный баланс
# ============================================================
# Push: 2 (1V + 1H) | Pull: 2 (1V + 1H) | Rear delt: 1
def day3_upper(phase):
    return fill_day(phase, [
        ("Сила (основное)", "Подтягивания с весом", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Вертикальная тяга, мёртвый вис до подбородка"),
        ("Сила (основное)", "Жим лэндмэйна стоя (одна рука)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Вертикальный жим, корпус стабилен"),
        ("Сила (основное)", "Жим штанги лёжа", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Горизонтальный жим, полная амплитуда"),
        ("Сила (основное)", "Тяга штанги в наклоне", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Горизонтальная тяга, лопатки вместе"),
        ("Вспомогательные", "Отведение плеч в сторону (трос)", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Средняя дельта, больше времени под нагрузкой"),
        ("Вспомогательные", "Face pull (кроссовер)", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Задняя дельта + ротация плеча"),
    ])


# ============================================================
# DAY 4: LEGS + ARMS
# ============================================================
def day4_legs_arms(phase):
    return fill_day(phase, [
        ("Сила (основное)", "Приседания с трэп грифом (широкая стойка)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Акцент на приводящие + ягодицы"),
        ("Сила (основное)", "Выпады шагающие (штанга на спине)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Со штангой, баланс + стабильность"),
        ("Сила (основное)", "Румынская тяга (штанга)", "main_sets", "main_reps", "main_weight", "main_rpe", "main_rest", "Мягкие колени, растяжение задней поверхности"),
        ("Вспомогательные", "Ягодичный мостик (одна нога)", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "С блином или гантелью, контроль"),
        ("Вспомогательные", "Разгибание ног сидя", "acc_sets", "acc_reps", "", "acc_rpe", "acc_rest", "Квадрицепс, пиковое сокращение"),
        ("Руки", "Жим книзу (трицепс, кабель)", "arm_sets", "arm_reps", "", "arm_rpe", "arm_rest", "V-гриф, локти прижаты"),
        ("Руки", "Сгибания рук с гантелями сидя на наклонной", "arm_sets", "arm_reps", "", "arm_rpe", "arm_rest", "Растянутый бицепс в стартовой позиции"),
    ])


# ============================================================
# DELOAD ADJUSTMENT
# ============================================================
def apply_deload(exercises):
    out = []
    for ex in exercises:
        block, name, sets, reps, weight, rpe, rest, notes = ex
        if block in ("Разминка", "Заминка"):
            out.append(ex)
            continue
        if sets and sets != "":
            try:
                s = int(sets)
                s = max(2, s - 2)
                sets = str(s)
            except ValueError:
                pass
        if weight and weight != "":
            weight = weight.replace("85%", "65%").replace("82%", "65%").replace("80%", "60%").replace("75%", "60%").replace("70%", "55%")
        if rpe and rpe != "":
            rpe = "6"
        if rest and rest != "":
            rest = rest.replace("180", "240").replace("120", "180").replace("90", "120").replace("60", "90")
        notes = (notes + " — РАЗГРУЗКА").strip()
        out.append((block, name, sets, reps, weight, rpe, rest, notes))
    return out


# ============================================================
# WORKBOOK GENERATION
# ============================================================
PHASE_MAP = {1: 1, 2: 1, 3: 1, 4: 1, 5: 2, 6: 2, 7: 2, 8: 2, 9: 3, 10: 3, 11: 3, 12: 3}
WEEKLY_FOCUS = {
    1: "Гипертрофия: адаптация, объём, техника",
    2: "Гипертрофия: прогрессия объёма",
    3: "Гипертрофия: пиковый объём",
    4: "РАЗГРУЗКА: -30% объёма, техника",
    5: "Сила: базовая сила, низкие повторы",
    6: "Сила: рост интенсивности",
    7: "Сила: пиковая интенсивность",
    8: "РАЗГРУЗКА: -30% объёма, техника",
    9: "Сила + Гипертрофия: переходный блок",
    10: "Сила + Гипертрофия: рост нагрузки",
    11: "Сила + Гипертрофия: пик цикла",
    12: "Сила + Гипертрофия: завершение цикла",
}


def write_day(ws, start_row, day_name, exercises, phase_num):
    row = start_row
    is_deload = phase_num in (4, 8)
    if is_deload:
        exercises = apply_deload(exercises)

    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row, 1, day_name)
    ws.cell(row, 1).font = Font(bold=True, size=12, color=WHITE)
    ws.cell(row, 1).fill = fill(NAVY)
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    for c, h in enumerate(HEADERS, 1):
        ws.cell(row, c, h)
    style_header(ws, row, 1, len(HEADERS))
    r0 = row
    row += 1

    for ex in exercises:
        for c, val in enumerate(ex, 1):
            if c <= len(HEADERS) and val is not None and val != "":
                ws.cell(row, c, val)
        row += 1

    for r in range(r0 + 1, row):
        section = str(ws.cell(r, 1).value or "")
        if section in ("Разминка", "Заминка"):
            for c in range(1, 9):
                ws.cell(r, c).fill = fill(LIGHT_BLUE)
        elif section == "Сила (основное)":
            for c in range(1, 9):
                ws.cell(r, c).fill = fill(BLUE)
        elif section == "Вспомогательные":
            for c in range(1, 9):
                ws.cell(r, c).fill = fill(GREEN)
        elif section == "Руки":
            for c in range(1, 9):
                ws.cell(r, c).fill = fill(YELLOW)
        elif section == "Гонка":
            for c in range(1, 9):
                ws.cell(r, c).fill = fill(ORANGE)

    ws.merge_cells(start_row=r0, start_column=9, end_row=row - 1, end_column=9)
    ws.cell(r0, 9).value = "Выполнение"
    ws.cell(r0, 9).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)

    return row + 1


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    # ============ OVERVIEW ============
    ov = wb.create_sheet("Обзор")
    ov.merge_cells("A1:H1")
    style_title(ov, "A1", "Сила + Гипертрофия: 12-недельный макроцикл")

    info = [
        ("Цель", "Максимальная гипертрофия + силовой прогресс. 4 тренировки в неделю: верх/ноги+руки."),
        ("Формат", "День 1: Верх A (гориз. жим + гориз. тяга + верт. жим + верт. тяга) — День 2: Ноги + Руки — День 3: Верх B (верт. тяга + верт. жим + гориз. жим + гориз. тяга) — День 4: Ноги + Руки"),
        ("Баланс", "Каждый день верха: 2 жима + 2 тяги, 2 вертикальных + 2 горизонтальных движения"),
        ("Фаза 1 (нед 1-3)", "Гипертрофия: 8-12 повторов, 70-75%, RPE 7-8, отдых 60-90 с"),
        ("Фаза 2 (нед 5-7)", "Сила: 4-6 повторов, 80-85%, RPE 8-9, отдых 180 с"),
        ("Фаза 3 (нед 9-12)", "Смешанная: 6-8 повторов, 75-82%, RPE 8-9, отдых 90-120 с"),
        ("Разгрузка", "Недели 4 и 8: -30% объёма, RPE 6, техника. Обязательна."),
    ]
    for i, (lbl, val) in enumerate(info):
        r = 3 + i
        ov.cell(r, 1, lbl).font = Font(bold=True)
        ov.cell(r, 1).fill = fill(LIGHT_GRAY)
        ov.cell(r, 2, val)
        ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        for c in range(1, 9):
            ov.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
            ov.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)

    # Must-include exercises list
    ov.merge_cells("A10:H10")
    style_title(ov, "A10", "Ключевые упражнения программы")
    key_ex = [
        "1. Жим лэндмэйна стоя",
        "2. Приседания с трэп грифом",
        "3. Тяга горизонтальная одной рукой",
        "4. Подтягивания с весом",
        "5. Выпады шагающие",
        "6. Жим гантелей на наклонной скамье",
        "7. Отведение плеч (латеральные подъёмы)",
        "8. Ягодичный мостик",
        "9. Вертикальная тяга одной рукой",
        "10. Отжимания на брусьях с весом",
    ]
    for i, ex in enumerate(key_ex):
        r = 11 + i
        ov.cell(r, 1, ex)
        ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ov.cell(r, 1).font = Font(size=10)
        ov.cell(r, 1).alignment = Alignment(vertical="center")

    for w in range(1, 9):
        ov.column_dimensions[get_column_letter(w)].width = 18 if w < 3 else 12
    set_common(ov)

    # ============ WEEKLY SHEETS ============
    day_generators = {
        1: ("День 1 — Верх A: Жим H + Тяга H + Жим V + Тяга V", day1_upper),
        2: ("День 2 — Ноги + Руки", day2_legs_arms),
        3: ("День 3 — Верх B: Тяга V + Жим V + Жим H + Тяга H", day3_upper),
        4: ("День 4 — Ноги + Руки", day4_legs_arms),
    }

    for week_num in range(1, 13):
        ws = wb.create_sheet(f"Неделя {week_num}")
        ws.merge_cells("A1:I1")
        phase = PHASE_MAP[week_num]
        phase_name = PHASES[phase]["name"]
        subtitle = f"Фаза: {phase_name} | {WEEKLY_FOCUS[week_num]}"
        if week_num in (4, 8):
            subtitle = f"⚠️ {subtitle}"
        style_title(ws, "A1", f"Неделя {week_num}: {subtitle}")
        ws.row_dimensions[1].height = 28

        row = 3
        for day_id in range(1, 5):
            day_name, gen_fn = day_generators[day_id]
            exercises = gen_fn(phase)
            row = write_day(ws, row, day_name, exercises, week_num)
            row += 1

        widths = [16, 40, 10, 14, 20, 14, 12, 38, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        set_common(ws)

    # ============ NOTES ============
    notes_sheet = wb.create_sheet("Примечания")
    notes_sheet.merge_cells("A1:C1")
    style_title(notes_sheet, "A1", "Примечания к программе")
    notes = [
        ("Баланс движений", "Каждый день верха: 2 жима (1 гориз. + 1 верт.) и 2 тяги (1 гориз. + 1 верт.) + задняя дельта. Полный баланс плоскостей."),
        ("Трэп-гриф", "Приседания с трэп грифом: ноги на ширине плеч, спина прямая, гриф поднимается за счёт ног и ягодиц. Альтернатива: румынская тяга с трэп грифом."),
        ("Жим лэндмэйна", "Стоя, одна рука, разноимённая нога впереди. Корпус стабилен, жим диагонально вверх. Нагрузка на переднюю дельту + грудь."),
        ("Жим штанги лёжа", "Заменил жим Арнольда. Классический горизонтальный жим для силового прогресса и гипертрофии груди."),
        ("Тяга одной рукой", "Корпус параллельно полу, лопатка тянется к позвоночнику. Без ротации корпуса."),
        ("Подтягивания с весом", "Полная амплитуда: мёртвый вис → подбородок выше грифа. Вес добавлять при 6+ повторах."),
        ("Брусья с весом", "Вертикальный жим. Локти вдоль корпуса (акцент на трицепс) или разведены (акцент на грудь)."),
        ("Отведение плеч", "Лёгкий вес, 12-15 повторов. Корпус без раскачки, мизинец вверх для среднего пучка."),
        ("Ягодичный мостик", "Лопатки на скамье, штанга на бёдрах. Максимальное сокращение в верхней точке на 1-2 сек."),
        ("Прогрессия силы", "Если выполнены все подходы в верхней границе повторов при RPE ≤ целевому — +2.5-5% на следующей неделе."),
        ("Прогрессия гипертрофии", "Если выполнены все подходы в верхней границе повторов — добавить 1 подход или +2.5 кг."),
        ("Разгрузка", "Недели 4 и 8: снижение подходов на 2, веса на 15-20%, RPE 6. Отдых увеличьте. Сохраняйте технику."),
        ("Сон/Питание", "Сон 7.5-8 ч. Белок 1.6-2 г/кг веса. Углеводы до/после тренировки. Вода 2-3 л/день."),
    ]
    for i, (t, txt) in enumerate(notes, 3):
        notes_sheet.cell(i, 1, t).font = Font(bold=True)
        notes_sheet.cell(i, 1).fill = fill(LIGHT_GRAY)
        notes_sheet.cell(i, 2, txt)
        notes_sheet.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        for c in range(1, 4):
            notes_sheet.cell(i, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
            notes_sheet.cell(i, c).alignment = Alignment(vertical="center", wrap_text=True)
    notes_sheet.column_dimensions["A"].width = 22
    notes_sheet.column_dimensions["B"].width = 60
    notes_sheet.column_dimensions["C"].width = 50
    set_common(notes_sheet)

    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    create_workbook()
