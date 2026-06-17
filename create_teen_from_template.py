import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

SRC = "/Users/iuriisho/OpenCode/Шаблон_hyrox_ведение.xlsx"
DST = "/Users/iuriisho/OpenCode/Подросток_12_недель.xlsx"

shutil.copy2(SRC, DST)

wb = openpyxl.load_workbook(DST)

NAVY = "FF1F4E78"
NAVY_DARK = "FF17365D"
LIGHT_BLUE = "FFEAF4FB"
YELLOW = "FFFFF2CC"
GRAY = "FFE7E6E6"
WHITE = "FFFFFFFF"
BLACK = "FF000000"

thin_border = Border(
    left=Side(style="thin", color="FFD9D9D9"),
    right=Side(style="thin", color="FFD9D9D9"),
    top=Side(style="thin", color="FFD9D9D9"),
    bottom=Side(style="thin", color="FFD9D9D9"),
)

# Day config: 3 days per week
DAYS = [
    ("Вторник", "A — Фулл-боди"),
    ("Четверг", "B — Фулл-боди"),
    ("Суббота", "C — Фулл-боди"),
]

# Phase config
PHASES = {
    1: {"sets": 2, "reps": "12–15", "weight": "1–2 кг", "tempo": "2-0-1-0", "rest": "60 с", "rpe": "7", "label": "Фаза 1 — Техника"},
    2: {"sets": 3, "reps": "12–15", "weight": "2–3 кг", "tempo": "2-0-1-0", "rest": "45–60 с", "rpe": "7–8", "label": "Фаза 2 — Развитие"},
    3: {"sets": 3, "reps": "10–12", "weight": "2–4.5 кг", "tempo": "3-0-1-0", "rest": "45–60 с", "rpe": "7–8", "label": "Фаза 3 — Укрепление"},
}

PHASE_MAP = {1: 1, 2: 1, 3: 1, 4: 1, 5: 2, 6: 2, 7: 2, 8: 2, 9: 3, 10: 3, 11: 3, 12: 3}

WEEKLY_FOCUS = {
    1: "Освоение техники, лёгкий вес",
    2: "Закрепление формы",
    3: "Уверенное выполнение",
    4: "Контроль + стабильность",
    5: "Увеличение объёма: +1 подход",
    6: "Рост нагрузки",
    7: "Удержание качества",
    8: "Повышение интенсивности",
    9: "Укрепление + контроль",
    10: "Пик формы",
    11: "Стабильность под нагрузкой",
    12: "Завершение цикла",
}

base_exercises = {
    "A": [
        ("Основные", "Гоблет-приседания с гантелью", "sets", "reps", "weight", "tempo", "rest", "rpe", "Колени по стопам, спина прямая"),
        ("Основные", "Выпады в стороны с гантелью", "sets", "reps", "weight", "tempo", "rest", "rpe", "Носок наружу, колено по стопе"),
        ("Основные", "Отжимания узким хватом (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Локти вдоль корпуса"),
        ("Основные", "Тяга гантели одной рукой (в наклоне)", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Спина прямо, локоть назад"),
        ("Основные", "Жим гантелей стоя", "sets", "reps", "weight", "tempo", "rest", "rpe", "Не прогибаться в пояснице"),
        ("Кор", "Планка Копенгагена", "sets", "15–20 с на сторону", "вес тела", "—", "30 с", "rpe", "Верхняя нога на возвышении"),
    ],
    "B": [
        ("Основные", "Выпады назад с гантелями", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Колено передней ноги за носком"),
        ("Основные", "Сиси-приседания (у стены/с опорой)", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени назад, пятки можно поднять"),
        ("Основные", "Отжимания от пола (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Корпус прямая линия"),
        ("Основные", "Тяга гантели к поясу (в упоре на колено)", "sets×2", "reps", "weight", "tempo", "rest", "rpe", "Локоть выше корпуса"),
        ("Основные", "Нордические наклоны (с резинкой/руками)", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени прижаты, спина прямо"),
        ("Кор", "Ротации с резинкой (дровосек)", "sets×2", "reps", "резинка", "—", "rest", "rpe", "Скручивание корпусом, руки прямые"),
    ],
    "C": [
        ("Основные", "Ягодичный мост с гантелью на бёдрах", "sets", "reps", "weight", "tempo", "rest", "rpe", "Сокращение 1 сек вверху"),
        ("Основные", "Обратные нордические (с рук/резинки)", "sets", "reps", "вес тела", "—", "rest", "rpe", "Колени на полу, корпус назад"),
        ("Основные", "Отжимания узким хватом (с колен/полные)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Локти вдоль корпуса"),
        ("Основные", "Боковые подъёмы гантелей", "sets", "reps", "weight", "tempo", "rest", "rpe", "Мизинец вверх, без рывков"),
        ("Основные", "Тяга резинки к лицу (Face Pull)", "sets", "reps", "резинка", "—", "rest", "rpe", "Локти в стороны, лопатки вместе"),
        ("Кор", "Подъёмы ног лёжа (низ живота)", "sets", "reps", "вес тела", "tempo", "rest", "rpe", "Поясница прижата к полу"),
    ],
}

TOTAL_COLS = 13
PLAN_COLS = 8

# helpers
def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def navy_dark_fill():
    return PatternFill("solid", fgColor=NAVY_DARK)

def light_blue_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE)

def yellow_fill():
    return PatternFill("solid", fgColor=YELLOW)

def gray_fill():
    return PatternFill("solid", fgColor=GRAY)

def white_fill():
    return PatternFill("solid", fgColor=WHITE)

def fill_workout(day_key, phase_num):
    p = PHASES[phase_num]
    base = base_exercises[day_key]
    out = []
    for ex in base:
        block, name, sets_k, reps_k, weight_k, tempo_k, rest_k, rpe_k, note = ex
        s = str(p["sets"]) if sets_k == "sets" else f'{p["sets"]}×2'
        rep = str(p["reps"])
        w = str(p["weight"]) if weight_k == "weight" else weight_k
        t = str(p["tempo"]) if tempo_k == "tempo" else tempo_k
        r = str(p["rest"]) if rest_k == "rest" else rest_k
        rpe = str(p["rpe"]) if rpe_k == "rpe" else rpe_k
        out.append([block, name, s, rep, w, rpe, t, r, "", "", "", "", note])
    return out


def write_week_sheet(ws, week_num):
    phase_num = PHASE_MAP[week_num]
    p = PHASES[phase_num]

    # Clear all existing data in the sheet
    ws.delete_rows(1, ws.max_row)
    # Clear leftover merged cells (delete_rows doesn't fully remove them)
    ws.merged_cells.ranges.clear()

    # Set column widths (match template)
    widths = [12.25, 29.75, 9.63, 11.38, 15.75, 10.5, 12.25, 10.5, 9.63, 13.0, 13.0, 13.0, 22.75]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # Row 1: Title
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    c = ws.cell(r, 1, f"Подросток 12 лет | Неделя {week_num}")
    c.font = Font(name="Arial", bold=True, size=15, color=WHITE)
    c.fill = navy_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    # Row 2: Week focus
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    c = ws.cell(r, 1, f"Период: Неделя {week_num} | {p['label']}")
    c.font = Font(name="Arial", size=9)
    c.fill = gray_fill()
    c.alignment = Alignment(horizontal="left", vertical="center")
    r += 1

    # Row 3: Fill note
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    c = ws.cell(r, 1, "Заполнять после каждого упражнения: факт подходов, повторов, веса, RPE и короткий комментарий.")
    c.font = Font(name="Arial", size=9)
    c.fill = yellow_fill()
    c.alignment = Alignment(horizontal="left", vertical="center")
    r += 1

    # Empty row
    r += 1

    for day_idx, (day_name, day_type) in enumerate(DAYS):
        day_key = ["A", "B", "C"][day_idx]

        # Day header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
        c = ws.cell(r, 1, f"{day_name} | {day_type}")
        c.font = Font(name="Arial", bold=True, size=12, color=WHITE)
        c.fill = navy_dark_fill()
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 21.75
        r += 1

        # Day focus
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
        c = ws.cell(r, 1, f"🎯 Фокус: {WEEKLY_FOCUS[week_num]}")
        c.font = Font(name="Arial", size=9)
        c.fill = yellow_fill()
        c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

        # Table headers
        headers = ["Раздел", "Упражнение", "План подходы", "План повторы", "План вес", "RPE/RIR", "Темп", "Отдых", "Факт подходы", "Факт повторы", "Факт вес", "Факт RPE", "Комментарий клиента"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(r, i, h)
            c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
            c.fill = navy_fill()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
        r += 1

        # Разминка
        warmup_data = ["Разминка", "Суставная разминка + лёгкий бег на месте", "5–7 мин"]
        for i, val in enumerate(warmup_data, 1):
            c = ws.cell(r, i, val)
            c.font = Font(name="Arial", size=9)
            c.fill = light_blue_fill()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border
        r += 1

        # Exercises
        exercises = fill_workout(day_key, phase_num)
        for ex in exercises:
            for i, val in enumerate(ex, 1):
                if val:
                    c = ws.cell(r, i, val)
                else:
                    c = ws.cell(r, i, "")
                c.font = Font(name="Arial", size=9)
                c.fill = light_blue_fill()
                c.alignment = Alignment(vertical="center", wrap_text=True)
                c.border = thin_border
            r += 1

        # Заминка
        cooldown_data = ["Заминка", "Растяжка: квадрицепс, задняя поверхность, кобра, поза ребёнка", "5 мин"]
        for i, val in enumerate(cooldown_data, 1):
            c = ws.cell(r, i, val)
            c.font = Font(name="Arial", size=9)
            c.fill = light_blue_fill()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border
        r += 1

        # Empty separator between days
        r += 1

    ws.print_options.fitToWidth = 1
    ws.page_setup.fitToWidth = 1


# Process all W1-W12 sheets
for week_num in range(1, 13):
    sheet_name = f"W{week_num}"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        write_week_sheet(ws, week_num)
        print(f"  Updated {sheet_name}")

# Update the Профиль клиента sheet with teen-specific info
if "Профиль клиента" in wb.sheetnames:
    ws = wb["Профиль клиента"]
    # Update info fields
    for row in ws.iter_rows(min_row=1, max_row=20):
        for c in row:
            if c.value and isinstance(c.value, str):
                if "HYROX" in c.value:
                    c.value = c.value.replace("HYROX", "Подросток 12 лет")
    print("  Updated Профиль клиента")

wb.save(DST)
print(f"\nСохранено: {DST}")
