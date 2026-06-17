import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

CYCLE_NAME = "Базовый старт"
TOTAL_WEEKS = 4
DAYS = ["Понедельник", "Вторник", "Четверг", "Пятница"]
DAY_TYPES = ["A — Фулл-боди (базовый)", "B — Низ + кор", "C — Фулл-боди (интенсив)", "D — Верх + кардио"]

PLAN_HEADERS = ["Блок", "Упражнение", "Подходы", "Повторы", "Вес", "Темп", "Отдых", "RPE", "Заметки"]
COL_WIDTHS = [18, 48, 10, 12, 10, 14, 10, 10, 36]
ALL_HEADERS_COUNT = len(PLAN_HEADERS)

NAVY = "1F3864"
NAVY_DARK = "2E5090"
LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
DAY_COLORS = {
    "A": "DAEEF3",
    "B": "E2EFDA",
    "C": "FCE4D6",
    "D": "EDEDED",
}

thin = Side(style="thin", color="999999")

workouts = {
    "A": [
        ["Основные", "Приседания с гантелями (гоблет)", 3, "15–20", "3 кг", "3-0-1-0", "60 с", "7–8", ""],
        ["Основные", "Жим гантелей лёжа (на коврике)", 3, "12–15", "3 кг", "3-0-1-0", "60 с", "7–8", ""],
        ["Основные", "Тяга гантели в наклоне (одна рука)", "3×2", "12–15", "3 кг", "2-0-1-0", "45 с", "7–8", ""],
        ["Доп", "Боковые подъёмы гантелей", 3, "15–20", "3 кг", "3-0-1-0", "45 с", "7–8", ""],
        ["Доп", "Ягодичный мостик с гантелью на бёдрах", 3, 20, "3 кг", "2-0-1-0", "45 с", "7", ""],
        ["Кор", "Планка на прямых руках", 3, "30–45 с", "—", "—", "30 с", "7–8", ""],
        ["Кор", "Подъёмы ног лёжа (низ живота)", 3, "15–20", "—", "2-0-1-0", "30 с", "7–8", ""],
    ],
    "B": [
        ["Низ", "Выпады назад с гантелями", "3×2", "12–15", "3 кг", "3-0-1-0", "60 с", "7–8", ""],
        ["Низ", "Румынская тяга на одной ноге (гантель)", "3×2", "12–15", "3 кг", "3-0-0-1", "60 с", "7–8", ""],
        ["Низ", "Резинка — боковые шаги (мини-банда)", 3, "15–20 на ногу", "резинка", "—", "45 с", "7", ""],
        ["Низ", "Резинка — ягодичный мостик с разведением", 3, 20, "резинка", "2-0-1-0", "45 с", "7–8", ""],
        ["Кор", "Скручивания с гантелью к груди", 3, 20, "3 кг", "2-0-1-0", "30 с", "7–8", ""],
        ["Кор", "Русский твист с гантелью", 3, "20 (10+10)", "3 кг", "2-0-1-0", "30 с", "7", ""],
        ["Кор", "Берёзка / обратная планка", 2, "30–45 с", "—", "—", "30 с", "7", ""],
    ],
    "C": [
        ["Комбо", "Приседания + жим над головой", 3, 15, "3 кг", "2-0-1-0", "60 с", "8", ""],
        ["Основные", "Отжимания от пола (ноги на диван / колени)", 3, "до упора", "собств. вес", "3-0-1-0", "60 с", "8–9", ""],
        ["Основные", "Тяга гантели к поясу (в упоре на колено/диван)", "3×2", 15, "3 кг", "2-0-1-0", "45 с", "8", ""],
        ["Основные", "Степ-ап на стул/диван с гантелями", "3×2", "12–15", "3 кг", "2-0-1-0", "45 с", "7–8", ""],
        ["Руки", "Сгибание рук с гантелями (бицепс)", 3, "15–20", "3 кг", "3-0-1-0", "30 с", "8", ""],
        ["Руки", "Разгибание руки с гантелью из-за головы (трицепс)", "3×2", 15, "3 кг", "3-0-1-0", "30 с", "8", ""],
        ["Кор", "V-складка (V-up)", 3, "12–15", "—", "2-0-1-0", "30 с", "8", ""],
    ],
    "D": [
        ["Верх", "Жим гантелей сидя (на стуле/диване)", 3, 15, "3 кг", "3-0-1-0", "45 с", "8", ""],
        ["Верх", "Пуловер с гантелью лёжа на диване", 3, 15, "3 кг", "3-0-1-0", "45 с", "7–8", ""],
        ["Верх", "Разводка гантелей лёжа", 3, 15, "3 кг", "3-0-1-0", "45 с", "7–8", ""],
        ["Верх", "Резинка — разведение рук (задняя дельта)", 3, 20, "резинка", "2-0-1-0", "30 с", "7–8", ""],
        ["Кардио", "Берпи без прыжка", 3, "10–12", "—", "—", "30 с", "8–9", ""],
        ["Кардио", "Махи гантелью (свинг)", 3, 20, "3 кг", "—", "30 с", "8", ""],
        ["Кор", "Планка с касанием плеч", 3, "20 (10+10)", "—", "—", "30 с", "7–8", ""],
    ],
}

warmup = [
    ["Круговые движения плечами", "30 с × 2"],
    ["Махи ногами вперёд/назад", "15 × 2"],
    ["Наклоны в стороны", "15 × 2"],
    ["Глубокие приседания без веса", "15"],
    ["Резинка — разведение рук назад", "15"],
]

cooldown = [
    ["Растяжка квадрицепса стоя", "30 с × 2"],
    ["Наклон к ногам сидя", "45 с"],
    ["Кобра / лёгкий прогиб", "30 с"],
    ["Скручивание лёжа (поза ребёнка)", "45 с"],
]

progression = [
    ["1", "Базовые повторения", "Обычный", "60–45 с", "7"],
    ["2", "+2–3 повтора к базе", "Эксцентрик 3 с", "45 с", "7–8"],
    ["3", "+2–3 повтора к W2", "3-0-1-0", "30–45 с", "8"],
    ["4", "Финальный push (макс повторы)", "3-0-1-0 (добавить 4-й сет)", "30 с", "8–9 / отказ"],
]


def fmt(row):
    return Font(size=9)


def header_font(bold=True, size=9, color=WHITE):
    return Font(bold=bold, size=size, color=color)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def style_cell(cell, font_obj=None, fill_obj=None, alignment_obj=None):
    if font_obj:
        cell.font = font_obj
    if fill_obj:
        cell.fill = fill_obj
    if alignment_obj:
        cell.alignment = alignment_obj
    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def write_title(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(bold=True, size=14, color=WHITE), fill(NAVY),
               Alignment(horizontal="left", vertical="center"))


def write_subtitle(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(size=10), fill(YELLOW),
               Alignment(horizontal="left", vertical="center"))
    cell.font = Font(size=10)


def write_day_header(ws, row, text, total_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    cell = ws.cell(row, 1, text)
    style_cell(cell, Font(bold=True, size=11, color=WHITE), fill(NAVY_DARK),
               Alignment(horizontal="left", vertical="center"))


def write_table_header(ws, row, headers, fill_color=NAVY):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        style_cell(cell, Font(bold=True, size=9, color=WHITE), fill(fill_color),
                   Alignment(horizontal="center", vertical="center", wrap_text=True))


def write_data_row(ws, row, values, color=LIGHT_BLUE):
    for c in range(1, ALL_HEADERS_COUNT + 1):
        cell = ws.cell(row, c, values[c - 1] if c <= len(values) and values[c - 1] is not None and values[c - 1] != "" else "")
        style_cell(cell, Font(size=9), fill(color),
                   Alignment(vertical="center", wrap_text=True))


def write_simple_table(ws, start_row, headers, rows, header_color=NAVY, row_color=LIGHT_BLUE):
    write_table_header(ws, start_row, headers, header_color)
    for i, row_data in enumerate(rows):
        write_data_row(ws, start_row + 1 + i, row_data, row_color)
    return start_row + 1 + len(rows)


wb = openpyxl.Workbook()

for week_num in range(1, TOTAL_WEEKS + 1):
    ws = wb.create_sheet(f"W{week_num}")
    total_cols = ALL_HEADERS_COUNT

    write_title(ws, 1, f"{CYCLE_NAME} | Неделя {week_num}", total_cols)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    cell = ws.cell(2, 1, "Заполняйте после каждого упражнения: факт подходов, повторов, веса, RPE, комментарий.")
    style_cell(cell, Font(size=10), fill(YELLOW),
               Alignment(horizontal="left", vertical="center"))

    r = 4

    for day_idx, day_name in enumerate(DAYS):
        day_key = ["A", "B", "C", "D"][day_idx]
        day_label = f"{day_name} | {DAY_TYPES[day_idx]}"
        day_color = DAY_COLORS[day_key]

        write_day_header(ws, r, day_label, total_cols)
        r += 1

        write_table_header(ws, r, PLAN_HEADERS)
        r += 1

        for ex in workouts[day_key]:
            write_data_row(ws, r, ex, day_color)
            r += 1

        r += 1

    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c - 1]

    ws.print_options.fitToWidth = 1
    ws.page_setup.fitToWidth = 1

# --- Sheet: Инфо ---
ws_info = wb.create_sheet("Инфо", 0)
total_cols = 4
for c in range(1, 5):
    ws_info.column_dimensions[get_column_letter(c)].width = [28, 50, 16, 24][c - 1]

write_title(ws_info, 1, f"{CYCLE_NAME} | Домашние тренировки", total_cols)

r = 3
ws_info.cell(r, 1, "Инвентарь").font = Font(bold=True, size=11)
r += 1
items = ["Гантели 2 × 3 кг", "Кольцевые резиночки (мини-банды)", "Коврик"]
for item in items:
    ws_info.cell(r, 1, f"• {item}")
    ws_info.cell(r, 1).font = Font(size=10)
    r += 1

r += 2
ws_info.cell(r, 1, "Структура недели").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["День", "Тип тренировки"],
                   [["Пн", "A — Фулл-боди (базовый)"],
                    ["Вт", "B — Низ + кор"],
                    ["Чт", "C — Фулл-боди (интенсив)"],
                    ["Пт", "D — Верх + кардио"]],
                   NAVY, LIGHT_BLUE)

r += 7
ws_info.cell(r, 1, "Прогрессия по неделям").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Неделя", "Повторения", "Темп", "Отдых", "RPE цель"],
                   progression, NAVY_DARK, DAY_COLORS["A"])

r += 8
ws_info.cell(r, 1, "Разминка (перед каждой тренировкой)").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Упражнение", "Длительность"],
                   warmup, NAVY, LIGHT_BLUE)

r += 6
ws_info.cell(r, 1, "Заминка").font = Font(bold=True, size=11)
r += 1
write_simple_table(ws_info, r, ["Упражнение", "Длительность"],
                   cooldown, NAVY_DARK, LIGHT_BLUE)

wb.remove(wb["Sheet"])

filename = f"/Users/iuriisho/OpenCode/Базовый_старт_домашняя_тренировка.xlsx"
wb.save(filename)
print(f"Saved: {filename}")
